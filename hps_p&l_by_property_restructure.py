import re
import streamlit as st
import pandas as pd

st.set_page_config(page_title="HPS Tools", layout="wide")


### ── HELPERS ─────────────────────────────────────────────────────────────────

def convert_letter_to_index(letter):
    index = 0
    for char in letter.upper().strip():
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def get_account_label(row, label_columns):
    label = ""
    for col in label_columns:
        val = row[col]
        if pd.notna(val) and str(val).strip():
            label = str(val).strip()
    return label


def clean_owner_name(owner):
    if pd.isna(owner):
        return ""
    return " ".join(str(owner).replace("(", "").replace(")", "").split())


def normalize_property_name(name):
    if pd.isna(name):
        return name
    text = str(name).strip()
    text = re.sub(r'(\s+\.\d+|\s+X+)+$', '', text, flags=re.IGNORECASE).strip()
    return text


def apply_property_merges(df, merges):
    rename_map = {}
    for group in merges:
        canonical = group["canonical"]
        for variant in group["variants"]:
            if variant != canonical:
                rename_map[variant] = canonical
    if not rename_map:
        return df
    df = df.copy()
    df["Property"] = df["Property"].replace(rename_map)
    owner_lookup = df.groupby("Property")["Owner"].first()
    has_dept = "_group_by_dept" in df.columns and bool(df["_group_by_dept"].any())
    group_keys = ["Accounting Period", "Account", "Property", "Department"] if has_dept else ["Accounting Period", "Account", "Property"]
    numeric_cols = [c for c in df.select_dtypes(include="number").columns if c != "_group_by_dept"]
    df = df.groupby(group_keys, as_index=False)[numeric_cols].sum()
    df["Owner"] = df["Property"].map(owner_lookup)
    df["_group_by_dept"] = has_dept
    return df


ROLLUP_ACCOUNTS = {"Gross Profit", "Net Ordinary Income", "Net Other Income"}


def is_rollup_account(account):
    if pd.isna(account):
        return True
    text = str(account).strip()
    return text == "" or text.lower().startswith("total ") or text in ROLLUP_ACCOUNTS


def process_file(df, first_data_col_idx, period_month_end):
    label_columns = df.columns[:first_data_col_idx]
    header_property = df.iloc[0]
    header_owner = df.iloc[1]
    body = df.iloc[2:].copy()

    body["Account"] = body.apply(
        lambda r: get_account_label(r, label_columns), axis=1
    )

    records = []
    n_cols = len(df.columns)

    for _, row in body.iterrows():
        account = str(row["Account"]).strip()
        if not account:
            continue
        for col_idx in range(first_data_col_idx, n_cols):
            prop_name = header_property.iloc[col_idx]
            prop_owner = header_owner.iloc[col_idx]
            prop_owner_text = "" if pd.isna(prop_owner) else str(prop_owner).strip()
            is_grand_total = col_idx == n_cols - 1
            is_owner_subtotal = "total" in prop_owner_text.lower() and not is_grand_total
            raw_prop_name = "N/A" if pd.isna(prop_name) else str(prop_name).strip()
            records.append({
                "Account": account,
                "Amount": row.iloc[col_idx],
                "Property Name": normalize_property_name(raw_prop_name),
                "Property Owner": prop_owner_text,
                "Is_Owner_Subtotal": is_owner_subtotal,
                "Is_Grand_Total": is_grand_total,
            })

    df_long = pd.DataFrame(records)
    df_long["Property Owner"] = df_long["Property Owner"].apply(clean_owner_name)

    valid = df_long.groupby("Account")["Amount"].apply(lambda s: s.notna().any())
    df_long = df_long[df_long["Account"].isin(valid[valid].index)].copy()
    df_long = df_long[~df_long["Account"].apply(is_rollup_account)].copy()
    df_long = df_long[
        (df_long["Is_Owner_Subtotal"] == False) | (df_long["Is_Grand_Total"] == True)
    ].copy()

    df_long["Accounting Period"] = period_month_end.date()
    return df_long


TOOL3_DEFAULT_ACCOUNT_MAP = {
    "Booking Income": "Booking Income",
    "Cleaning Fees": "Cleaning Income",
    "Credit Card Income": "Credit Card Income",
    "Damage Income": "Damage Income",
    "Linen Program Fee Income": "Linen Program Fee",
    "Management Fee Income": "Management Fee Income",
    "Markup - Appliance": "Markups",
    "Markup - Cleaning": "Markups",
    "Markup - Furniture": "Markups",
    "Markup - Guest Expenses": "Markups",
    "Markup - Licenses and Permits": "Markups",
    "Markup - Locks": "Markups",
    "Markup - Materials": "Markups",
    "Markup - Photos": "Markups",
    "Markup - Postage": "Markups",
    "Markup - Repair Labor": "Markups",
    "Markup - Staging": "Markups",
    "Markup - Unit Inventory": "Markups",
    "Markups": "Markups",
    "Airbnb Booking Fee": "Other Revenue",
    "Booking.com Commission": "Other Revenue",
    "Break Fee Income": "Other Revenue",
    "Cancellation Insurance": "Other Revenue",
    "Chargeback": "Other Revenue",
    "Early/Late Fee Check In/Out": "Other Revenue",
    "Expedia Commission": "Other Revenue",
    "Float Fees Forfeited": "Other Revenue",
    "Home Away Commission": "Other Revenue",
    "Lease Application Fee": "Other Revenue",
    "Occupancy Violation": "Other Revenue",
    "Pool Heating": "Other Revenue",
    "Referral Income": "Other Revenue",
    "Refund Cancellation": "Other Revenue",
    "Reimbursed Income": "Other Revenue",
    "Rentals United Commission": "Other Revenue",
    "Storage Locker Rental Income": "Other Revenue",
    "Trip Insurance Income": "Other Revenue",
    "Pet Fees": "Pet Fees",
    "Rent": "Rental Income",
    "Credit Card Fees": "Credit Card Fees",
    "Guest Expenses": "Other COGS",
    "Linen Program Fee": "Other COGS",
    "Owners Reimbursement": "Other COGS",
    "OTA Fees": "OTA Fees",
    "Owners Proceed": "Owners Rental Proceeds",
    "Health Insurance": "Insurance",
    "Property-Liability Insurance": "Insurance",
    "Trip Insurance Expenses": "Insurance",
    "Licenses and Permits": "Licenses and Permits",
    "Management Fee Adjustment": "Management Fee Adjustment",
    "Management Fees": "Management Fee Expense",
    "Bad debt": "Other G&A",
    "Bank Service Charges": "Other G&A",
    "Break Fee": "Other G&A",
    "Delivery": "Other G&A",
    "Lease Application Fees": "Other G&A",
    "Not Collected Sales Tax": "Other G&A",
    "Office Supplies": "Other G&A",
    "Penalty Expense": "Other G&A",
    "Postage & Delivery": "Other G&A",
    "Rent Expenses": "Other G&A",
    "Accounting Fees": "Professional Fees",
    "Consulting": "Professional Fees",
    "Legal Fees": "Professional Fees",
    "Professional Fees": "Professional Fees",
    "VA Subcontractor": "Professional Fees",
    "Cable/Internet": "Telephone & Utilities",
    "Electricity & Heat": "Telephone & Utilities",
    "Fax": "Telephone & Utilities",
    "Gas Utility": "Telephone & Utilities",
    "Phone": "Telephone & Utilities",
    "Water & Sewer": "Telephone & Utilities",
    "Entertainment": "Travel & Entertainment",
    "Meals": "Travel & Entertainment",
    "Travel": "Travel & Entertainment",
    "Paid Time Off": "Payroll Costs",
    "Parental Leave Time Off": "Payroll Costs",
    "Payroll Accounting": "Payroll Costs",
    "Payroll Bonus": "Payroll Costs",
    "Payroll Clearing": "Payroll Costs",
    "Payroll Fees": "Payroll Costs",
    "Payroll Guest Services": "Payroll Costs",
    "Payroll Inter Company": "Payroll Costs",
    "Payroll Marketing": "Payroll Costs",
    "Payroll Mgmt": "Payroll Costs",
    "Payroll Operations": "Payroll Costs",
    "Payroll Overtime": "Payroll Costs",
    "Payroll Owner Services": "Payroll Costs",
    "Payroll Taxes": "Payroll Costs",
    "Payroll Vacation": "Payroll Costs",
    "Worker's Compensation": "Payroll Costs",
    "Conference": "Professional Development",
    "Training": "Professional Development",
    "Recruiting Expense": "Recruiting Expense",
    "Cleaning Inspector": "Cleaning",
    "Cleaning Supplies": "Cleaning",
    "Cleaning Units": "Cleaning",
    "Garage Cleaning": "Cleaning",
    "Laundry Attendant Payroll": "Cleaning",
    "Linen Inventory": "Cleaning",
    "Unit Inventory": "Cleaning",
    "AC Filters": "Maintenance",
    "Appliances": "Maintenance",
    "Auto Allowance": "Maintenance",
    "Consumables": "Maintenance",
    "Electric": "Maintenance",
    "Furniture & Decorations": "Maintenance",
    "Garage Maintenance": "Maintenance",
    "HVAC": "Maintenance",
    "HVAC Repairs": "Maintenance",
    "Landscape Expense": "Maintenance",
    "Materials": "Maintenance",
    "Moveable": "Maintenance",
    "Parking Lot": "Maintenance",
    "Plumbing": "Maintenance",
    "Repairs": "Maintenance",
    "Subcontractor": "Maintenance",
    "Trash Removal": "Maintenance",
    "Cleaning Slippage": "Slippage",
    "Inventory Slippage": "Slippage",
    "Linen Slippage": "Slippage",
    "Maintenance Slippage": "Slippage",
    "Equipment Rental": "Transportation & Equipment",
    "Gas & Maintenance": "Transportation & Equipment",
    "Truck Rental": "Transportation & Equipment",
    "Advertising": "Ad & Marketing",
    "Marketing": "Ad & Marketing",
    "Referral Bonus": "Sales Bonuses & Promotions",
    "Sign on Bonus": "Sales Bonuses & Promotions",
    "Staff Promotion": "Sales Bonuses & Promotions",
    "Staging Bonus": "Sales Bonuses & Promotions",
    "Unit Photos": "Other Marketing Expenses",
    "Website Fees": "Other Marketing Expenses",
    "Dues and Subscriptions": "Dues and Subscriptions",
    "Software": "Software",
    "Admin OH Expenses Split by Unit": "Other Expense",
    "Insurance Reimbursement": "Other Expense",
    "Interest Expense": "Other Expense",
    "R&M OH Expenses Split by Unit": "Other Expense",
    "Transfer FROM Float Fees": "Other Expense",
    "Transfer FROM Reserve Funds": "Other Expense",
    "Transfer FROM Unit": "Other Expense",
    "Write Off": "Other Expense",
    "Interest Income": "Other Income",
    "Other Income Drew": "Other Income",
    "Tax Collection Allowance": "Other Income",
    "Transfer TO Float Fees": "Other Income",
    "Transfer TO Reserve Fund": "Other Income",
    "Transfer TO Unit": "Other Income",
}

TOOL3_EXPENSE_CATEGORIES = sorted(set(TOOL3_DEFAULT_ACCOUNT_MAP.values()))

TOOL4_DEFAULT_DEPT_FILTER_ACCOUNTS = [
    "Paid Time Off",
    "Payroll Clearing",
    "Payroll Fees",
    "Payroll Inter Company",
    "Payroll Mgmt",
    "Payroll Overtime",
    "Payroll Taxes",
    "Payroll Vacation",
    "Worker's Compensation",
    "Parental Leave Time Off",
    "Cable/Internet",
    "Fax",
    "Phone",
    "Health Insurance",
    "Bonuses",
    "Dues and Subscriptions",
    "Software",
    "Accounting Fees",
    "Consulting",
    "Legal Fees",
    "Payroll Owner Services",
    "Professional Fees",
    "VA Subcontractor",
]

TOOL4_DEFAULT_DEPT_FILTER_DEPARTMENTS = [
    "Owner Success:Owner Sales",
    "Owner Success:Owner Service",
]


### ── TOOL 5 HELPERS ──────────────────────────────────────────────────────────

TOOL5_SECTION_LABELS = {"Income", "Cost of Goods Sold", "Expense", "Other Income", "Other Expense"}
TOOL5_INCOME_SECTIONS = {"Income", "Other Income"}
TOOL5_ROLLUP_ACCOUNTS = ROLLUP_ACCOUNTS | {"Net Income"}
TOOL5_TOLERANCE = 0.01


def tool5_is_rollup_account(account):
    if account is None:
        return True
    text = str(account).strip()
    return text == "" or text.lower().startswith("total ") or text in TOOL5_ROLLUP_ACCOUNTS


def tool5_find_first_data_col(header_row):
    for i, val in enumerate(header_row):
        if pd.notna(val) and str(val).strip():
            return i
    raise ValueError("Could not find a data column in the header row.")


def tool5_parse_pnl_sections(raw_df, body_start_row, first_data_col, total_col_idx, include_rollups=False):
    """Parse a QuickBooks-style P&L export (Portfolio or Property-level) into
    one row per leaf account, tagged with its top-level section (Income, Cost
    of Goods Sold, Expense, Other Income, Other Expense) and its value in the
    trailing TOTAL column. By default, subtotal/rollup rows ("Total ...",
    Gross Profit, Net Ordinary Income, Net Other Income, Net Income) are
    dropped, since callers doing reconciliation or GL-account matching only
    want leaf accounts. Pass include_rollups=True to keep those rows too
    (e.g. for a flat, faithful reproduction of the original report)."""
    label_columns = list(range(0, first_data_col))
    body = raw_df.iloc[body_start_row:]
    current_section = None
    records = []
    for _, row in body.iterrows():
        account = get_account_label(row, label_columns)
        if not account:
            continue
        if account in TOOL5_SECTION_LABELS:
            current_section = account
            continue
        data_values = row.iloc[first_data_col:total_col_idx + 1]
        if data_values.notna().sum() == 0:
            continue
        if tool5_is_rollup_account(account) and not include_rollups:
            continue
        records.append({"Section": current_section, "Account": account, "Total": row.iloc[total_col_idx]})
    return pd.DataFrame(records, columns=["Section", "Account", "Total"])


def tool5_build_consolidated_portfolio_pnl(portfolio_raw_list, period_labels):
    """Consolidate multiple periods' Portfolio P&L into one flat table: every
    row from the original report (leaf accounts AND subtotal/rollup rows like
    "Total Markup" or "Net Income") is kept, aligned across periods by Account
    name. An account that only exists in some periods still gets its own row,
    blank for the periods where it's absent. Row order is an ordered merge
    across periods: the first period sets the initial order, and any account
    introduced only in a later period is inserted right after the nearest
    account it follows in that period's own order (rather than appended at
    the very end), so a brand-new account still lands in its natural section
    instead of trailing after Net Income."""
    period_dfs = []
    for portfolio_raw in portfolio_raw_list:
        header = portfolio_raw.iloc[0]
        first_data_col = tool5_find_first_data_col(header)
        total_col = portfolio_raw.shape[1] - 1
        period_dfs.append(tool5_parse_pnl_sections(portfolio_raw, 1, first_data_col, total_col, include_rollups=True))

    row_order = []
    seen = set()
    section_lookup = {}
    for period_df in period_dfs:
        insert_at = 0
        for account, section in zip(period_df["Account"], period_df["Section"]):
            if account in seen:
                insert_at = row_order.index(account) + 1
            else:
                seen.add(account)
                row_order.insert(insert_at, account)
                section_lookup[account] = "" if tool5_is_rollup_account(account) else section
                insert_at += 1

    consolidated = pd.DataFrame({
        "Section": [section_lookup[a] for a in row_order],
        "Account": row_order,
        "Is Rollup": [tool5_is_rollup_account(a) for a in row_order],
    })

    for label, period_df in zip(period_labels, period_dfs):
        totals = period_df.set_index("Account")["Total"]
        consolidated[label] = consolidated["Account"].map(totals)

    return consolidated


def tool5_export_consolidated_pnl_excel(consolidated_df, period_labels):
    """Write the consolidated Portfolio P&L to an .xlsx with subtotal/rollup
    rows (Total ..., Gross Profit, Net Ordinary Income, Net Other Income, Net
    Income) bolded and top-bordered so they visually stand apart from leaf
    accounts, the way a printed financial statement would."""
    import io
    from openpyxl.styles import Font, Border, Side

    export_df = consolidated_df[["Section", "Account"] + list(period_labels)]
    n_cols = export_df.shape[1]
    first_period_col = 3  # 1-indexed: Section=1, Account=2, periods start at 3

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Consolidated P&L")
        worksheet = writer.sheets["Consolidated P&L"]

        bold_font = Font(bold=True)
        top_border = Border(top=Side(style="thin"))
        number_format = "#,##0.00;(#,##0.00)"

        for row_offset, is_rollup in enumerate(consolidated_df["Is Rollup"]):
            excel_row = row_offset + 2
            for col in range(first_period_col, n_cols + 1):
                worksheet.cell(row=excel_row, column=col).number_format = number_format
            if is_rollup:
                for col in range(1, n_cols + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.font = bold_font
                    cell.border = top_border

        for col_idx, column in enumerate(export_df.columns, start=1):
            max_len = max([len(str(column))] + [len(str(v)) for v in export_df[column].fillna("")])
            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    return buffer.getvalue()


def tool5_attach(base_df, source_df, value_col, dup_accounts):
    """Left-merge value_col from source_df onto base_df, keyed by Account for
    accounts that are unambiguous, and by (Account, Section) for accounts that
    appear under more than one section in the Portfolio P&L (e.g. an account
    used as both an Income and an Expense line)."""
    nondup_source = source_df[~source_df["Account"].isin(dup_accounts)][["Account", value_col]]
    dup_source = source_df[source_df["Account"].isin(dup_accounts)][["Account", "Section", value_col]]
    base_nondup = base_df[~base_df["Account"].isin(dup_accounts)].merge(nondup_source, on="Account", how="left")
    base_dup = base_df[base_df["Account"].isin(dup_accounts)].merge(dup_source, on=["Account", "Section"], how="left")
    return pd.concat([base_nondup, base_dup], ignore_index=True)


def tool5_reconcile(portfolio_raw, property_raw, gl_raw):
    """3-way reconcile the Portfolio P&L, Property-level P&L, and GL transaction
    detail for the same accounting period. Returns (recon_df, extra_gl_df)."""

    # Portfolio P&L: row 0 is the month/TOTAL header, accounts start at row 1.
    portfolio_header = portfolio_raw.iloc[0]
    p_first_data_col = tool5_find_first_data_col(portfolio_header)
    p_total_col = portfolio_raw.shape[1] - 1
    if str(portfolio_header.iloc[p_total_col]).strip().upper() != "TOTAL":
        raise ValueError("Portfolio P&L: last column is not labeled TOTAL.")
    portfolio_df = tool5_parse_pnl_sections(portfolio_raw, 1, p_first_data_col, p_total_col)
    portfolio_df = portfolio_df.rename(columns={"Total": "Portfolio Total"})

    # Property-level P&L: row 0 is property names, row 1 is owner/TOTAL header,
    # accounts start at row 2.
    property_name_header = property_raw.iloc[0]
    property_owner_header = property_raw.iloc[1]
    pr_first_data_col = tool5_find_first_data_col(property_name_header)
    pr_total_col = property_raw.shape[1] - 1
    if str(property_owner_header.iloc[pr_total_col]).strip().upper() != "TOTAL":
        raise ValueError("Property-Level P&L: last column is not labeled TOTAL.")
    property_df = tool5_parse_pnl_sections(property_raw, 2, pr_first_data_col, pr_total_col)
    property_df = property_df.rename(columns={"Total": "Property-Level Total"})

    # GL transaction detail: header is row 0 already (real column names).
    gl = gl_raw.copy()
    gl.columns = gl.columns.astype(str).str.strip()
    required_gl_cols = {"Type", "Account", "Class", "Amount", "Item"}
    missing_cols = required_gl_cols - set(gl.columns)
    if missing_cols:
        raise ValueError(f"GL file is missing required column(s): {', '.join(sorted(missing_cols))}")

    gl_real = gl[gl["Type"].notna() & gl["Account"].notna() & (gl["Class"] == "SICB Management")].copy()
    gl_real["Item Prefix"] = gl_real["Item"].astype(str).str.split(":").str[0]
    gl_real["Section"] = gl_real["Item Prefix"].apply(lambda p: "Income" if p == "Income" else "Expense")

    portfolio_accounts = set(portfolio_df["Account"])

    # GL accounts with no match on the Portfolio P&L are Balance Sheet accounts
    # by definition (a P&L report can only ever contain Income Statement
    # accounts), so they're surfaced separately rather than flagged as issues.
    gl_extra = gl_real[~gl_real["Account"].isin(portfolio_accounts)]
    gl_extra_grouped = (
        gl_extra.groupby("Account", as_index=False)["Amount"].sum()
        .rename(columns={"Amount": "GL Total"})
    )
    gl_extra_grouped = gl_extra_grouped.reindex(
        gl_extra_grouped["GL Total"].abs().sort_values(ascending=False).index
    ).reset_index(drop=True)

    gl_matched = gl_real[gl_real["Account"].isin(portfolio_accounts)]

    section_counts = portfolio_df.groupby("Account")["Section"].nunique()
    dup_accounts = set(section_counts[section_counts > 1].index)

    # Unambiguous accounts: sum every matching GL row regardless of the
    # item-derived section (avoids losing dollars to inconsistent Item tagging).
    gl_nondup = (
        gl_matched[~gl_matched["Account"].isin(dup_accounts)]
        .groupby("Account", as_index=False)["Amount"].sum()
    )
    # Ambiguous accounts (same leaf name under more than one section, e.g.
    # "Trip Insurance" as both Income and Expense): split by the GL Item prefix.
    gl_dup = (
        gl_matched[gl_matched["Account"].isin(dup_accounts)]
        .groupby(["Account", "Section"], as_index=False)["Amount"].sum()
    )

    recon_nondup = portfolio_df[~portfolio_df["Account"].isin(dup_accounts)].merge(
        gl_nondup, on="Account", how="left"
    )
    recon_dup = portfolio_df[portfolio_df["Account"].isin(dup_accounts)].merge(
        gl_dup, on=["Account", "Section"], how="left"
    )
    recon = pd.concat([recon_nondup, recon_dup], ignore_index=True)
    recon = recon.rename(columns={"Amount": "GL Total"})
    recon["GL Total"] = recon["GL Total"].fillna(0.0)

    property_present = set(zip(property_df["Account"], property_df["Section"]))
    recon = tool5_attach(recon, property_df, "Property-Level Total", dup_accounts)
    recon["Property-Level Total"] = recon["Property-Level Total"].fillna(0.0)
    recon["Portfolio Total"] = recon["Portfolio Total"].fillna(0.0)

    recon["GL Total (Adjusted)"] = recon.apply(
        lambda r: -r["GL Total"] if r["Section"] in TOOL5_INCOME_SECTIONS else r["GL Total"], axis=1
    )

    def _checker(r):
        diff_property = r["Portfolio Total"] - r["Property-Level Total"]
        diff_gl = r["Portfolio Total"] - r["GL Total (Adjusted)"]
        property_ok = abs(diff_property) < TOOL5_TOLERANCE
        gl_ok = abs(diff_gl) < TOOL5_TOLERANCE
        if property_ok and gl_ok:
            return "Match"
        parts = []
        if not property_ok:
            if (r["Account"], r["Section"]) not in property_present:
                parts.append("not broken out at property level (expected)")
            else:
                parts.append(f"vs Property {diff_property:,.2f}")
        if not gl_ok:
            parts.append(f"vs GL {diff_gl:,.2f}")
        return "Mismatch: " + "; ".join(parts)

    recon["Checker"] = recon.apply(_checker, axis=1)
    recon = recon[["Section", "Account", "Portfolio Total", "Property-Level Total", "GL Total (Adjusted)", "Checker"]]
    recon = recon.sort_values(["Section", "Account"]).reset_index(drop=True)

    return recon, gl_extra_grouped


TOOL5_OWNED_ENTITIES = {"CBTS LP", "CIF LP", "PBC LP", "KES LP", "CBC LP", "CinCB LP", "SinCB LP"}


def tool5_parse_period_pnls(portfolio_raw, property_raw):
    """Parse a single period's Portfolio and Property-level P&L into their
    per-account, per-section DataFrames (see tool5_parse_pnl_sections)."""
    portfolio_header = portfolio_raw.iloc[0]
    p_first_data_col = tool5_find_first_data_col(portfolio_header)
    p_total_col = portfolio_raw.shape[1] - 1
    portfolio_df = tool5_parse_pnl_sections(portfolio_raw, 1, p_first_data_col, p_total_col)

    property_name_header = property_raw.iloc[0]
    pr_first_data_col = tool5_find_first_data_col(property_name_header)
    pr_total_col = property_raw.shape[1] - 1
    property_df = tool5_parse_pnl_sections(property_raw, 2, pr_first_data_col, pr_total_col)

    return portfolio_df, property_df


def tool5_build_account_universe(portfolio_df_list, property_df_list):
    """Combine every uploaded period's parsed Portfolio/Property P&L into one
    global account universe, section lookup, and duplicate-section account
    map. This must be computed across ALL periods together, not per-period —
    otherwise an account that's only ambiguous in some periods (e.g. "Trip
    Insurance" showing under both Income and Expense in one year but only
    Income in another) would get split in one period's data but not
    another's, producing inconsistent account labels once periods are
    combined. The account universe itself must include accounts that net to
    zero at the Portfolio level (pure reallocation accounts, e.g. "Admin OH
    Expenses Split by Unit") but still carry real per-property detail on the
    Property-level P&L — anchoring on Portfolio alone would silently drop
    those accounts' GL transactions from the extraction."""
    account_section = {}
    dup_candidate_sections = {}
    expanded_accounts = set()

    for portfolio_df in portfolio_df_list:
        for account, section in zip(portfolio_df["Account"], portfolio_df["Section"]):
            account_section.setdefault(account, section)
            dup_candidate_sections.setdefault(account, set()).add(section)
        expanded_accounts |= set(portfolio_df["Account"])

    for property_df in property_df_list:
        for account, section in zip(property_df["Account"], property_df["Section"]):
            account_section.setdefault(account, section)
        expanded_accounts |= set(property_df["Account"])

    dup_account_sections = {a: s for a, s in dup_candidate_sections.items() if len(s) > 1}
    return expanded_accounts, account_section, dup_account_sections


def tool5_split_owner_property(name):
    text = str(name).strip()
    if ":" in text:
        owner, prop = text.split(":", 1)
        return owner.strip(), prop.strip()
    return "Corporate", "Corporate"


def tool5_clean_owner(owner):
    if owner == "Corporate":
        return owner
    return re.sub(r' -C$', '', owner).strip()


def tool5_owner_type(owner):
    if owner in TOOL5_OWNED_ENTITIES:
        return "Owned"
    if owner == "Corporate":
        return "Corporate"
    return "Third Party"


def tool5_default_department(department):
    if pd.isna(department):
        return department
    text = str(department).strip()
    if ":" in text:
        return text.split(":", 1)[1].strip()
    return text


TOOL5_RAW_GL_PASSTHROUGH_COLS = ["Date", "Type", "Num", "Class", "Name", "Source Name", "Item", "Item Description", "Split", "Memo"]


def tool5_extract_unit_economics(gl_raw, expanded_accounts, account_section, dup_account_sections):
    """Extract Owner/Property/Department per relevant GL transaction (Class =
    SICB Management, Account in the expanded P&L universe). Owner/Property come
    from splitting Name on the first ':' ("Owner -C:Property"); Names with no
    colon (e.g. "SICB - Rent", "SICB - Customer") are unattributed and go to a
    "Corporate" bucket for both Owner and Property."""
    gl = gl_raw.copy()
    gl.columns = gl.columns.astype(str).str.strip()
    required_cols = {"Type", "Account", "Class", "Name", "Amount", "Date"}
    missing_cols = required_cols - set(gl.columns)
    if missing_cols:
        raise ValueError(f"GL file is missing required column(s): {', '.join(sorted(missing_cols))}")

    gl_real = gl[
        gl["Type"].notna() & gl["Account"].notna() & (gl["Class"] == "SICB Management")
        & gl["Account"].isin(expanded_accounts)
    ].copy()

    for col in TOOL5_RAW_GL_PASSTHROUGH_COLS:
        if col not in gl_real.columns:
            gl_real[col] = pd.NA

    split_result = gl_real["Name"].apply(tool5_split_owner_property)
    gl_real["Owner"] = split_result.apply(lambda t: tool5_clean_owner(t[0]))
    gl_real["Property"] = split_result.apply(lambda t: t[1] if t[1] == "Corporate" else normalize_property_name(t[1]))
    gl_real["Property Owner Type"] = gl_real["Owner"].apply(tool5_owner_type)
    gl_real["Section"] = gl_real["Account"].map(account_section)
    gl_real["Accounting Period"] = (pd.to_datetime(gl_real["Date"], errors="coerce") + pd.offsets.MonthEnd(0)).dt.date

    # Accounts that appear under more than one P&L section (e.g. "Trip
    # Insurance" is both Income and Expense) get disambiguated per-transaction
    # using the GL's own Item column, and split into distinct account labels
    # ("Trip Insurance Income" / "Trip Insurance Expenses") so each section's
    # activity is unambiguous in the output rather than colliding under one
    # shared account name.
    for account, sections in dup_account_sections.items():
        income_section = next((s for s in sections if s in TOOL5_INCOME_SECTIONS), None)
        other_section = next((s for s in sections if s not in TOOL5_INCOME_SECTIONS), None)
        mask = gl_real["Account"] == account
        item_prefix = gl_real.loc[mask, "Item"].astype(str).str.split(":").str[0]
        is_income = item_prefix == "Income"
        gl_real.loc[mask, "Section"] = is_income.map({True: income_section, False: other_section})
        gl_real.loc[mask, "Account"] = is_income.map({True: f"{account} Income", False: f"{account} Expenses"})

    if "Department" not in gl_real.columns:
        gl_real["Department"] = pd.NA
    gl_real["Department (Raw)"] = gl_real["Department"]
    gl_real["Department"] = gl_real["Department"].apply(tool5_default_department)

    # "Num" mixes real numbers (check #) and text (e.g. "Zelle") in the raw GL,
    # which breaks Streamlit's Arrow-based table display if left as-is.
    gl_real["Num"] = gl_real["Num"].apply(lambda v: "" if pd.isna(v) else str(v))

    columns = (
        ["Accounting Period", "Account", "Section", "Department", "Department (Raw)", "Property", "Owner", "Property Owner Type"]
        + TOOL5_RAW_GL_PASSTHROUGH_COLS
        + ["Amount"]
    )
    return gl_real[columns].reset_index(drop=True)


TOOL5_OH_SPLIT_ACCOUNTS = {"Admin OH Expenses Split by Unit", "R&M OH Expenses Split by Unit"}
TOOL5_BELOW_NET_ORDINARY_SECTIONS = {"Other Income", "Other Expense"}


def tool5_apply_corporate_filter(df):
    """Drop Corporate (unattributed) rows, except for accounts below Net
    Ordinary Income (Other Income / Other Expense) which have no sensible
    per-property attribution — but still exclude the OH-split accounts even
    there, since their Corporate rows are just the offsetting reclass entry
    for dollars that are already fully captured in the per-property rows."""
    keep_mask = (df["Property"] != "Corporate") | (
        df["Section"].isin(TOOL5_BELOW_NET_ORDINARY_SECTIONS) & ~df["Account"].isin(TOOL5_OH_SPLIT_ACCOUNTS)
    )
    return df[keep_mask].copy()


def tool5_apply_pnl_sign(df):
    """Flip every transaction's sign uniformly (Amount = -raw Amount), so a
    positive Amount always means "increases Net Income" and negative always
    means "decreases Net Income" — regardless of Section. This one identity
    holds for every account because the GL's raw Debit-minus-Credit sign is
    already recorded per transaction line, not per account category: a normal
    revenue credit flips to positive, but a refund/credit-memo against revenue
    (a debit) flips to negative, correctly staying negative; a normal expense
    debit flips to negative, but a reversal/credit against an expense flips to
    positive, correctly showing it improved Net Income. No Section lookup is
    needed, and summing this column for any subset of rows gives that
    subset's Net Income directly."""
    df = df.copy()
    df["Amount"] = -df["Amount"]
    return df


def tool5_build_export_grouped(df):
    """Export 1 — Property-Level P&L data, grouped by Accounting Period +
    Account + Property, summed Amount, with Owner and Property Owner Type
    attached (first value seen per Property)."""
    filtered = tool5_apply_pnl_sign(tool5_apply_corporate_filter(df))
    owner_lookup = filtered.groupby("Property")["Owner"].first()
    owner_type_lookup = filtered.groupby("Property")["Property Owner Type"].first()

    grouped = filtered.groupby(["Accounting Period", "Account", "Property"], as_index=False)["Amount"].sum()
    grouped["Owner"] = grouped["Property"].map(owner_lookup)
    grouped["Property Owner Type"] = grouped["Property"].map(owner_type_lookup)
    grouped = grouped[["Accounting Period", "Account", "Property", "Amount", "Owner", "Property Owner Type"]]
    return grouped.sort_values(["Accounting Period", "Account", "Property"]).reset_index(drop=True)


def tool5_build_export_detail(df):
    """Export 2 — same Corporate-exclusion filter as Export 1, but one row per
    raw transaction (no grouping), with full GL detail retained."""
    filtered = tool5_apply_pnl_sign(tool5_apply_corporate_filter(df))
    columns = (
        ["Accounting Period", "Date", "Type", "Num", "Account", "Section", "Class", "Department",
         "Property", "Owner", "Property Owner Type", "Name", "Source Name", "Item", "Item Description",
         "Split", "Memo", "Amount"]
    )
    return filtered[columns].sort_values(["Accounting Period", "Account", "Property"]).reset_index(drop=True)


def tool5_to_export_csv(df, text_columns):
    """Format a Tool 5 export dataframe to CSV bytes: Accounting Period as
    YYYY-MM-DD, Amount as a plain 4-decimal number, and Accounting Period plus
    the given text columns wrapped Excel-safe as ="..." so they aren't
    auto-reformatted (e.g. Accounting Period turning into a real Excel date)."""
    out = df.copy()
    out["Accounting Period"] = out["Accounting Period"].apply(lambda d: d.strftime("%Y-%m-%d") if pd.notna(d) else "")
    out["Amount"] = out["Amount"].apply(lambda v: f"{v:.4f}" if pd.notna(v) else "")
    for col in ["Accounting Period"] + list(text_columns):
        if col in out.columns:
            out[col] = '="' + out[col].astype(str).str.replace('"', '""') + '"'
    return out.to_csv(index=False).encode("utf-8")


def tool5_apply_property_merges(df, merges):
    """Rename Property variants to their canonical name. Unlike Tool 2's
    apply_property_merges, this does NOT re-group/sum — Tool 5's transaction
    detail (Export 2) must stay at the raw-transaction grain, so grouping is
    left to tool5_build_export_grouped at export time."""
    rename_map = {}
    for group in merges:
        canonical = group["canonical"]
        for variant in group["variants"]:
            if variant != canonical:
                rename_map[variant] = canonical
    if not rename_map:
        return df
    df = df.copy()
    df["Property"] = df["Property"].replace(rename_map)
    return df


def tool5_portfolio_net_income(portfolio_raw_list):
    """The combined Portfolio P&L Net Income (bottom-line row, TOTAL column),
    summed across every uploaded period's Portfolio P&L file."""
    total = 0.0
    for portfolio_raw in portfolio_raw_list:
        net_income_rows = portfolio_raw.index[portfolio_raw[0] == "Net Income"]
        if len(net_income_rows) == 0:
            raise ValueError("Portfolio P&L: could not find a 'Net Income' row.")
        total_col = portfolio_raw.shape[1] - 1
        val = portfolio_raw.iloc[net_income_rows[0], total_col]
        total += val if pd.notna(val) else 0.0
    return total


def tool5_official_property_net_income(property_raw_list, merges):
    """Extract each property's official Net Income from the Property-level P&L
    (the report's own bottom-line row), keyed by the same normalized property
    name used elsewhere, with the same merges applied so combined properties'
    official Net Income is summed together for a fair comparison. Summed across
    every uploaded period's Property-level P&L file."""
    rename_map = {}
    for group in merges:
        canonical = group["canonical"]
        for variant in group["variants"]:
            if variant != canonical:
                rename_map[variant] = canonical

    official = {}
    for property_raw in property_raw_list:
        name_row = property_raw.iloc[0]
        net_income_rows = property_raw.index[property_raw[0] == "Net Income"]
        if len(net_income_rows) == 0:
            raise ValueError("Property-Level P&L: could not find a 'Net Income' row.")
        net_income_row = property_raw.iloc[net_income_rows[0]]
        n_cols = property_raw.shape[1]

        for c in range(6, n_cols - 1):
            name = name_row.iloc[c]
            if pd.notna(name):
                clean_name = normalize_property_name(str(name).strip())
                clean_name = rename_map.get(clean_name, clean_name)
                val = net_income_row.iloc[c]
                official[clean_name] = official.get(clean_name, 0.0) + (val if pd.notna(val) else 0.0)
    return official


def tool5_build_net_income_check(unit_econ_df, property_raw_list, merges):
    """Per-property 'final reconciliation': compare each property's Net Income
    derived from the (merged, Corporate-filtered) extracted GL data — combined
    across all uploaded periods — against the Property-level P&L's own Net
    Income for that property."""
    merged_df = tool5_apply_property_merges(unit_econ_df, merges)
    filtered = tool5_apply_corporate_filter(merged_df)
    gl_net_income = -filtered.groupby("Property")["Amount"].sum()

    official = tool5_official_property_net_income(property_raw_list, merges)

    all_props = sorted(set(official) | set(gl_net_income.index) - {"Corporate"})
    rows = []
    for p in all_props:
        pnl_val = official.get(p, 0.0)
        gl_val = gl_net_income.get(p, 0.0)
        diff = pnl_val - gl_val
        rows.append({
            "Property": p,
            "Property P&L Net Income": pnl_val,
            "GL Net Income": gl_val,
            "Diff": diff,
            "Match": abs(diff) < TOOL5_TOLERANCE,
        })
    return pd.DataFrame(rows)


### ── SESSION STATE INIT ──────────────────────────────────────────────────────

if "tool" not in st.session_state:
    st.session_state.tool = None
if "step" not in st.session_state:
    st.session_state.step = "upload"
if "accumulated" not in st.session_state:
    st.session_state.accumulated = pd.DataFrame()
if "raw_df" not in st.session_state:
    st.session_state.raw_df = None
if "tool2_step" not in st.session_state:
    st.session_state.tool2_step = "upload"
if "tool2_accumulated" not in st.session_state:
    st.session_state.tool2_accumulated = pd.DataFrame()
if "tool2_merges" not in st.session_state:
    st.session_state.tool2_merges = []
if "tool2_group_by_dept" not in st.session_state:
    st.session_state.tool2_group_by_dept = False
if "tool2_owner_type_map" not in st.session_state:
    st.session_state.tool2_owner_type_map = None  # None = skipped, dict = applied
if "tool2_raw_amount_sum" not in st.session_state:
    st.session_state.tool2_raw_amount_sum = 0.0
if "tool2_raw_transactions" not in st.session_state:
    st.session_state.tool2_raw_transactions = pd.DataFrame()
if "tool3_step" not in st.session_state:
    st.session_state.tool3_step = "upload"
if "tool3_accumulated" not in st.session_state:
    st.session_state.tool3_accumulated = pd.DataFrame()
if "tool3_merges" not in st.session_state:
    st.session_state.tool3_merges = []
if "tool3_group_by_dept" not in st.session_state:
    st.session_state.tool3_group_by_dept = False
if "tool3_owner_type_map" not in st.session_state:
    st.session_state.tool3_owner_type_map = None
if "tool3_raw_amount_sum" not in st.session_state:
    st.session_state.tool3_raw_amount_sum = 0.0
if "tool3_account_map" not in st.session_state:
    st.session_state.tool3_account_map = None
if "tool4_step" not in st.session_state:
    st.session_state.tool4_step = "upload"
if "tool4_expenses_df" not in st.session_state:
    st.session_state.tool4_expenses_df = pd.DataFrame()
if "tool4_units_df" not in st.session_state:
    st.session_state.tool4_units_df = pd.DataFrame()
if "tool4_result_df" not in st.session_state:
    st.session_state.tool4_result_df = pd.DataFrame()
if "tool4_account_mode_map" not in st.session_state:
    st.session_state.tool4_account_mode_map = {}
if "tool4_unit_owner_overrides" not in st.session_state:
    st.session_state.tool4_unit_owner_overrides = {}
if "tool4_dept_filter_accounts" not in st.session_state:
    st.session_state.tool4_dept_filter_accounts = list(TOOL4_DEFAULT_DEPT_FILTER_ACCOUNTS)
if "tool4_dept_filter_departments" not in st.session_state:
    st.session_state.tool4_dept_filter_departments = list(TOOL4_DEFAULT_DEPT_FILTER_DEPARTMENTS)
if "tool4_missing_overrides" not in st.session_state:
    st.session_state.tool4_missing_overrides = {}
if "tool4_dropped_df" not in st.session_state:
    st.session_state.tool4_dropped_df = pd.DataFrame()
if "tool5_step" not in st.session_state:
    st.session_state.tool5_step = "upload"
if "tool5_recon_df" not in st.session_state:
    st.session_state.tool5_recon_df = pd.DataFrame()
if "tool5_extra_gl_df" not in st.session_state:
    st.session_state.tool5_extra_gl_df = pd.DataFrame()
if "tool5_portfolio_raw_list" not in st.session_state:
    st.session_state.tool5_portfolio_raw_list = []
if "tool5_property_raw_list" not in st.session_state:
    st.session_state.tool5_property_raw_list = []
if "tool5_gl_raw_list" not in st.session_state:
    st.session_state.tool5_gl_raw_list = []
if "tool5_unit_econ_raw_df" not in st.session_state:
    st.session_state.tool5_unit_econ_raw_df = pd.DataFrame()
if "tool5_dept_remap" not in st.session_state:
    st.session_state.tool5_dept_remap = None
if "tool5_unit_econ_df" not in st.session_state:
    st.session_state.tool5_unit_econ_df = pd.DataFrame()
if "tool5_property_merges" not in st.session_state:
    st.session_state.tool5_property_merges = []
if "tool5_typo_overrides" not in st.session_state:
    st.session_state.tool5_typo_overrides = {}
if "tool5_net_income_check_df" not in st.session_state:
    st.session_state.tool5_net_income_check_df = pd.DataFrame()


### ── MENU ────────────────────────────────────────────────────────────────────

def go_home():
    st.session_state.tool = None
    st.session_state.step = "upload"
    st.session_state.accumulated = pd.DataFrame()
    st.session_state.raw_df = None
    st.session_state.tool2_step = "upload"
    st.session_state.tool2_accumulated = pd.DataFrame()
    st.session_state.tool2_merges = []
    st.session_state.tool2_group_by_dept = False
    st.session_state.tool2_owner_type_map = None
    st.session_state.tool2_raw_amount_sum = 0.0
    st.session_state.tool2_raw_transactions = pd.DataFrame()
    st.session_state.tool3_step = "upload"
    st.session_state.tool3_accumulated = pd.DataFrame()
    st.session_state.tool3_merges = []
    st.session_state.tool3_group_by_dept = False
    st.session_state.tool3_owner_type_map = None
    st.session_state.tool3_raw_amount_sum = 0.0
    st.session_state.tool3_account_map = None
    st.session_state.tool4_step = "upload"
    st.session_state.tool4_expenses_df = pd.DataFrame()
    st.session_state.tool4_units_df = pd.DataFrame()
    st.session_state.tool4_result_df = pd.DataFrame()
    st.session_state.tool4_account_mode_map = {}
    st.session_state.tool4_unit_owner_overrides = {}
    st.session_state.tool4_dept_filter_accounts = list(TOOL4_DEFAULT_DEPT_FILTER_ACCOUNTS)
    st.session_state.tool4_dept_filter_departments = list(TOOL4_DEFAULT_DEPT_FILTER_DEPARTMENTS)
    st.session_state.tool4_missing_overrides = {}
    st.session_state.tool4_dropped_df = pd.DataFrame()
    st.session_state.tool5_step = "upload"
    st.session_state.tool5_recon_df = pd.DataFrame()
    st.session_state.tool5_extra_gl_df = pd.DataFrame()
    st.session_state.tool5_portfolio_raw_list = []
    st.session_state.tool5_property_raw_list = []
    st.session_state.tool5_gl_raw_list = []
    st.session_state.tool5_unit_econ_raw_df = pd.DataFrame()
    st.session_state.tool5_dept_remap = None
    st.session_state.tool5_unit_econ_df = pd.DataFrame()
    st.session_state.tool5_property_merges = []
    st.session_state.tool5_typo_overrides = {}
    st.session_state.tool5_net_income_check_df = pd.DataFrame()


if st.session_state.tool is None:
    st.title("HPS Tools")
    st.write("Select a tool to get started.")
    st.divider()

    TOOLS = [
        {
            "key": "open_pnl",
            "tool": "pnl_restructure",
            "title": "P&L by Property",
            "subtitle": "Based on Monthly P&L by Property",
            "description": "Upload monthly P&L by Property Excel files, process one or more months, and export in long or wide CSV format.",
        },
        {
            "key": "open_tool2",
            "tool": "tool2",
            "title": "P&L by Property",
            "subtitle": "Based on GL Report",
            "description": "Upload GL Report Excel files, merge properties, map owner types, and export a clean flat CSV with reconciliation check.",
        },
        {
            "key": "open_tool3",
            "tool": "tool3",
            "title": "Company Expenses",
            "subtitle": "Data Prep",
            "description": "Upload GL expense files, map accounts to expense categories, rename departments, and export with period and category filters.",
        },
        {
            "key": "open_tool4",
            "tool": "tool4",
            "title": "SICB Management",
            "subtitle": "Expense Allocation",
            "description": "Allocate SICB Management expenses to individual properties based on active unit count per month.",
        },
        {
            "key": "open_tool5",
            "tool": "tool5",
            "title": "Final Property Level P&L",
            "subtitle": "3-Way Reconciliation",
            "description": "Upload the Portfolio P&L, Property-level P&L, and GL transaction detail for a period and confirm every account ties out across all three.",
        },
    ]

    card_style = """
        <style>
        div[data-testid="stVerticalBlock"] div[data-testid="stVerticalBlockBorderWrapper"] {
            height: 100%;
        }
        </style>
    """
    st.markdown(card_style, unsafe_allow_html=True)

    row1 = st.columns(3, gap="large")
    row2_cols = st.columns(3, gap="large")
    row2 = [row2_cols[0], row2_cols[1]]

    all_slots = row1 + row2

    for i, tool in enumerate(TOOLS):
        with all_slots[i]:
            with st.container(border=True):
                st.markdown(f"#### {tool['title']}")
                st.caption(tool["subtitle"])
                st.write(tool["description"])
                if st.button("Open", key=tool["key"], use_container_width=True, type="primary"):
                    st.session_state.tool = tool["tool"]
                    st.rerun()


### ── TOOL 1: P&L BY PROPERTY RESTRUCTURE ────────────────────────────────────

elif st.session_state.tool == "pnl_restructure":

    st.title("P&L by Property Data - Based on Monthly P&L by Property")
    if st.button("← Back to Menu", key="back_pnl"):
        go_home()
        st.rerun()

    st.divider()

    # ── STEP: UPLOAD ──────────────────────────────────────────────────────────

    if st.session_state.step == "upload":

        if not st.session_state.accumulated.empty:
            n = st.session_state.accumulated["Accounting Period"].nunique()
            st.info(f"{n} month(s) already loaded. Upload the next file.")

        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])

        if uploaded_file is not None:
            st.session_state.raw_df = pd.read_excel(uploaded_file, header=None)
            st.session_state.step = "inputs"
            st.rerun()

    # ── STEP: INPUTS ──────────────────────────────────────────────────────────

    elif st.session_state.step == "inputs":

        st.subheader("File Preview")
        st.dataframe(st.session_state.raw_df.head(10), use_container_width=True)

        col1, col2 = st.columns(2)
        with col1:
            col_letter = st.text_input("First data column letter (e.g. H)")
        with col2:
            date_input = st.text_input("Accounting period date (MM/DD/YYYY)")

        if st.button("Process", type="primary"):
            if not col_letter:
                st.error("Enter the first data column letter.")
            elif not date_input:
                st.error("Enter the accounting period date.")
            else:
                try:
                    col_idx = convert_letter_to_index(col_letter)
                    period_date = pd.to_datetime(date_input, format="%m/%d/%Y")
                    period_month_end = period_date + pd.offsets.MonthEnd(0)

                    with st.spinner(f"Processing {period_month_end.strftime('%B %Y')}..."):
                        month_df = process_file(
                            st.session_state.raw_df, col_idx, period_month_end
                        )

                    st.session_state.accumulated = pd.concat(
                        [st.session_state.accumulated, month_df], ignore_index=True
                    )
                    st.session_state.raw_df = None
                    st.session_state.step = "action"
                    st.rerun()

                except ValueError as e:
                    st.error(f"Error: {e} — check your date format (MM/DD/YYYY).")

    # ── STEP: ACTION ──────────────────────────────────────────────────────────

    elif st.session_state.step == "action":

        acc = st.session_state.accumulated
        n_months = acc["Accounting Period"].nunique()

        st.success(f"{n_months} month(s) loaded — {len(acc):,} total rows")
        st.dataframe(acc, use_container_width=True)

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Add Another Month", use_container_width=True):
                st.session_state.step = "upload"
                st.rerun()
        with col2:
            if st.button("Export", type="primary", use_container_width=True):
                st.session_state.step = "export"
                st.rerun()

    # ── STEP: EXPORT ──────────────────────────────────────────────────────────

    elif st.session_state.step == "export":

        acc = st.session_state.accumulated
        n_months = acc["Accounting Period"].nunique()

        st.success(f"Ready to export — {len(acc):,} rows across {n_months} month(s)")

        st.subheader("Version 1 — Long Format")
        st.caption(
            "One row per Account + Property + Month combination. "
            "Best for filtering, pivot tables, and loading into databases or BI tools."
        )
        acc_long = acc[(acc["Amount"].notna()) & (acc["Amount"] != 0) & (acc["Is_Grand_Total"] == False)].copy()
        acc_long = acc_long.drop(columns=["Is_Owner_Subtotal", "Is_Grand_Total"])

        owner_lookup_long = acc_long.groupby("Property Name")["Property Owner"].first()
        acc_long["Property Owner"] = acc_long["Property Name"].map(owner_lookup_long)
        st.dataframe(acc_long, use_container_width=True)

        csv_long = acc_long.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download Long Format CSV",
            data=csv_long,
            file_name="hps_pnl_long.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )

        st.divider()

        st.subheader("Version 2 — Wide Format")
        st.caption(
            "One row per Account + Property combination, with one column per month sorted earliest to latest. "
            "Best for side-by-side month comparison and sharing as a report."
        )

        acc_wide = acc.copy()
        acc_wide["Accounting Period"] = acc_wide["Accounting Period"].astype(str).str.replace("-", "/")
        acc_wide["Amount"] = acc_wide["Amount"].replace(0, pd.NA)

        sorted_months = sorted(acc_wide["Accounting Period"].unique())

        owner_lookup = acc_wide.groupby("Property Name")["Property Owner"].first().reset_index()

        wide_df = acc_wide.pivot_table(
            index=["Account", "Property Name"],
            columns="Accounting Period",
            values="Amount",
            aggfunc="first",
        ).reset_index()

        wide_df.columns.name = None
        wide_df = wide_df[["Account", "Property Name"] + sorted_months]
        wide_df = wide_df.merge(owner_lookup, on="Property Name", how="left")
        wide_df = wide_df[["Account", "Property Name", "Property Owner"] + sorted_months]

        st.dataframe(wide_df, use_container_width=True)

        csv_wide = wide_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download Wide Format CSV",
            data=csv_wide,
            file_name="hps_pnl_wide.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )

        st.divider()

        if st.button("Restart", use_container_width=True):
            st.session_state.accumulated = pd.DataFrame()
            st.session_state.raw_df = None
            st.session_state.step = "upload"
            st.rerun()


### ── TOOL 2: P&L BY PROPERTY DATA - BASED ON GL REPORT ──────────────────────

elif st.session_state.tool == "tool2":

    import csv

    st.title("P&L by Property Data - Based on GL Report")
    if st.button("← Back to Menu", key="back_tool2"):
        go_home()
        st.rerun()

    st.divider()

    # ── STEP: UPLOAD ──────────────────────────────────────────────────────────

    if st.session_state.tool2_step == "upload":

        if not st.session_state.tool2_accumulated.empty:
            n = st.session_state.tool2_accumulated["Accounting Period"].nunique()
            st.info(f"{n} file(s) already loaded. Upload the next file.")

        if st.session_state.tool2_accumulated.empty:
            checked = st.checkbox("Group by Department", value=st.session_state.tool2_group_by_dept)
            st.session_state.tool2_group_by_dept = checked
        else:
            dept_label = "on" if st.session_state.tool2_group_by_dept else "off"
            st.info(f"Department grouping is **{dept_label}** (set on first upload).")

        group_by_dept = st.session_state.tool2_group_by_dept

        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])

        if uploaded_file is not None:
            with st.spinner("Processing..."):
                df = pd.read_excel(uploaded_file, header=None)

                # Remove first 3 rows, use 4th row as header
                df = df.iloc[3:].reset_index(drop=True)
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)

                # Remove first 4 columns
                df = df.iloc[:, 4:]

                # Remove rows where Account is empty
                df.columns = df.columns.str.strip()
                df = df[df["Account"].notna()]
                df = df[df["Account"] != ""]
                df = df.reset_index(drop=True)

                # Create Accounting Period from Date
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
                df["Accounting Period"] = df["Date"] + pd.offsets.MonthEnd(0)
                df = df[df["Accounting Period"].notna()].reset_index(drop=True)

                # Fill missing Department so groupby doesn't silently drop those rows
                if group_by_dept and "Department" in df.columns:
                    df["Department"] = df["Department"].fillna("Unassigned")

                # Create Owner and Property from Name / Class / Memo
                _OWNED_ENTITIES = ["CBTS LP", "CIF LP", "PBC LP", "KES LP", "CBC LP", "CinCB LP", "SinCB LP"]
                _MEMO_OWNER_RE = re.compile(
                    r'(' + '|'.join(re.escape(e) for e in _OWNED_ENTITIES) + r'):([A-Za-z]+\s+\d+)'
                )
                owners = []
                properties = []
                for _, row in df.iterrows():
                    name_value = str(row["Name"])
                    class_value = str(row["Class"])
                    memo_value = str(row["Memo"]) if "Memo" in df.columns and pd.notna(row.get("Memo")) else ""
                    memo_match = _MEMO_OWNER_RE.search(memo_value)
                    if ":" in name_value:
                        parts = name_value.split(":", 1)
                        owners.append(parts[0].strip())
                        properties.append(parts[1].strip())
                    elif memo_match:
                        owners.append(memo_match.group(1).strip())
                        properties.append(memo_match.group(2).strip())
                    elif ":" in class_value:
                        parts = class_value.split(":", 1)
                        owners.append(parts[0].strip())
                        properties.append(parts[1].strip())
                    else:
                        owners.append(class_value.strip())
                        properties.append(class_value.strip())

                df["Owner"] = owners
                df["Property"] = properties

                # Clean Owner and Property
                df["Owner"] = df["Owner"].str.replace(" -C$", "", regex=True).str.strip().str.replace(r"\s+", " ", regex=True)
                df["Property"] = df["Property"].str.replace("XXX", "", regex=False).str.replace(r"\.\d{2}", "", regex=True).str.strip().str.replace(r"\s+", " ", regex=True)

                # Force Amount to numeric in case it was read as text
                if "Amount" in df.columns:
                    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")

                # Capture raw amount sum before grouping for reconciliation
                st.session_state.tool2_raw_amount_sum += df["Amount"].sum(min_count=1) if "Amount" in df.columns else 0.0

                # Save pre-grouped transactions for audit export
                st.session_state.tool2_raw_transactions = pd.concat(
                    [st.session_state.tool2_raw_transactions, df.copy()], ignore_index=True
                )

                # Lock first-seen Owner per Property before grouping
                owner_lookup = df.groupby("Property")["Owner"].first()

                # Group by Accounting Period + Account + Property, sum numeric columns
                group_keys = ["Accounting Period", "Account", "Property", "Department"] if group_by_dept else ["Accounting Period", "Account", "Property"]
                numeric_cols = df.select_dtypes(include="number").columns.tolist()
                df = df.groupby(group_keys, as_index=False)[numeric_cols].sum()
                df["Owner"] = df["Property"].map(owner_lookup)
                df["_group_by_dept"] = group_by_dept

            st.session_state.tool2_accumulated = pd.concat(
                [st.session_state.tool2_accumulated, df], ignore_index=True
            )
            st.session_state.tool2_step = "action"
            st.rerun()

    # ── STEP: ACTION ──────────────────────────────────────────────────────────

    elif st.session_state.tool2_step == "action":

        acc = st.session_state.tool2_accumulated
        n_files = acc["Accounting Period"].nunique()

        st.success(f"{n_files} file(s) loaded — {len(acc):,} total rows")
        st.dataframe(acc, use_container_width=True)

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("Add Another File", use_container_width=True):
                st.session_state.tool2_step = "upload"
                st.rerun()
        with col2:
            if st.button("Merge Properties", use_container_width=True):
                st.session_state.tool2_step = "merge"
                st.rerun()
        with col3:
            if st.button("Continue →", type="primary", use_container_width=True):
                st.session_state.tool2_owner_type_map = None
                st.session_state.tool2_step = "owner_type"
                st.rerun()

    # ── STEP: MERGE ───────────────────────────────────────────────────────────

    elif st.session_state.tool2_step == "merge":

        st.subheader("Merge Properties")
        st.caption("Group property names that refer to the same property. Select all variants, then pick which name to keep.")

        all_properties = sorted(st.session_state.tool2_accumulated["Property"].unique().tolist())

        if st.session_state.tool2_merges:
            st.write("**Current merge groups:**")
            for i, group in enumerate(st.session_state.tool2_merges):
                col1, col2 = st.columns([6, 1])
                with col1:
                    variants_str = ", ".join(f"`{v}`" for v in group["variants"] if v != group["canonical"])
                    st.write(f"{variants_str} → **{group['canonical']}**")
                with col2:
                    if st.button("Remove", key=f"remove_merge_{i}"):
                        st.session_state.tool2_merges.pop(i)
                        st.rerun()
            st.divider()

        st.write("**Add a new merge group:**")
        n = len(st.session_state.tool2_merges)
        selected_variants = st.multiselect(
            "Select property names to merge",
            options=all_properties,
            key=f"merge_variants_{n}",
        )

        if len(selected_variants) >= 2:
            canonical = st.radio(
                "Which name to keep?",
                options=selected_variants,
                key=f"merge_canonical_{n}",
            )
            if st.button("Add Group", type="primary"):
                st.session_state.tool2_merges.append({
                    "variants": selected_variants,
                    "canonical": canonical,
                })
                st.rerun()

        st.divider()

        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True):
                st.session_state.tool2_step = "action"
                st.rerun()
        with col2:
            if st.button("Apply & Continue", type="primary", use_container_width=True):
                if st.session_state.tool2_merges:
                    st.session_state.tool2_accumulated = apply_property_merges(
                        st.session_state.tool2_accumulated,
                        st.session_state.tool2_merges,
                    )
                st.session_state.tool2_step = "action"
                st.rerun()

    # ── STEP: OWNER TYPE MAPPING ──────────────────────────────────────────────

    elif st.session_state.tool2_step == "owner_type":

        OWNED_DEFAULT = {"CBTS LP", "CIF LP", "PBC LP", "KES LP", "CBC LP", "CinCB LP", "SinCB LP"}
        MGMT_DEFAULT = {"SICB Management"}

        st.subheader("Map Property Owner Type")
        st.caption(
            "Assign each Owner to a category. Everything not listed below is mapped as **Third Party**."
        )

        all_owners = sorted(st.session_state.tool2_accumulated["Owner"].dropna().unique().tolist())

        # Build working map from session state or defaults
        if st.session_state.tool2_owner_type_map is None:
            working_map = {}
            for o in all_owners:
                if o in OWNED_DEFAULT:
                    working_map[o] = "Owned"
                elif o in MGMT_DEFAULT:
                    working_map[o] = "SICB Management"
                else:
                    working_map[o] = "Third Party"
        else:
            working_map = dict(st.session_state.tool2_owner_type_map)
            for o in all_owners:
                if o not in working_map:
                    working_map[o] = "Third Party"

        st.write("**Current mapping** (only Owned and SICB Management shown — all others are Third Party):")

        owned_owners = [o for o, t in working_map.items() if t == "Owned"]
        mgmt_owners = [o for o, t in working_map.items() if t == "SICB Management"]

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Owned**")
            for o in owned_owners:
                st.markdown(f"- `{o}`")
        with col2:
            st.markdown("**SICB Management**")
            for o in mgmt_owners:
                st.markdown(f"- `{o}`")

        st.divider()
        st.write("**Add or change a mapping:**")

        third_party_owners = [o for o in all_owners if working_map.get(o) == "Third Party"]

        if third_party_owners:
            col_a, col_b = st.columns(2)
            with col_a:
                owner_to_add = st.selectbox(
                    "Select an Owner (currently Third Party)",
                    options=third_party_owners,
                    key="owner_type_select",
                )
            with col_b:
                new_type = st.radio(
                    "Map to",
                    options=["Owned", "SICB Management"],
                    key="owner_type_radio",
                    horizontal=True,
                )
            if st.button("Add Mapping", type="primary"):
                working_map[owner_to_add] = new_type
                st.session_state.tool2_owner_type_map = working_map
                st.rerun()
        else:
            st.info("All owners are already mapped.")

        st.divider()
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("← Back", use_container_width=True):
                st.session_state.tool2_step = "action"
                st.rerun()
        with col2:
            if st.button("Skip (no owner type column)", use_container_width=True):
                st.session_state.tool2_owner_type_map = None
                st.session_state.tool2_step = "period_filter"
                st.rerun()
        with col3:
            if st.button("Apply & Continue →", type="primary", use_container_width=True):
                st.session_state.tool2_owner_type_map = working_map
                st.session_state.tool2_step = "period_filter"
                st.rerun()

    # ── STEP: PERIOD FILTER ───────────────────────────────────────────────────

    elif st.session_state.tool2_step == "period_filter":

        acc = st.session_state.tool2_accumulated
        all_periods = sorted(acc["Accounting Period"].dropna().unique().tolist(), key=lambda x: str(x))

        st.subheader("Filter Accounting Periods")
        st.caption("All periods are selected by default — click any period to remove it from the export.")

        if "tool2_period_multiselect" not in st.session_state or not st.session_state["tool2_period_multiselect"]:
            st.session_state["tool2_period_multiselect"] = all_periods

        col_sel, col_desel = st.columns(2)
        with col_sel:
            if st.button("Select All", key="tool2_period_select_all", use_container_width=True):
                st.session_state["tool2_period_multiselect"] = all_periods
                st.rerun()
        with col_desel:
            if st.button("Deselect All", key="tool2_period_desel_all", use_container_width=True):
                st.session_state["tool2_period_multiselect"] = []
                st.rerun()

        selected_periods = st.multiselect(
            "Accounting periods to include",
            options=all_periods,
            key="tool2_period_multiselect",
            format_func=lambda x: str(x),
        )

        n_sel = len(selected_periods)
        n_tot = len(all_periods)
        if n_sel == 0:
            st.warning("No periods selected — select at least one to continue.")
        else:
            st.info(f"{n_sel} of {n_tot} period(s) selected.")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool2_period_back"):
                st.session_state.tool2_step = "owner_type"
                st.rerun()
        with col2:
            if st.button("Continue →", type="primary", use_container_width=True, key="tool2_period_continue", disabled=(n_sel == 0)):
                st.session_state.tool2_step = "export"
                st.rerun()

    # ── STEP: EXPORT ──────────────────────────────────────────────────────────

    elif st.session_state.tool2_step == "export":

        acc = st.session_state.tool2_accumulated.copy()

        # Apply period filter
        selected_periods = st.session_state.get("tool2_period_multiselect", [])
        if selected_periods:
            acc = acc[acc["Accounting Period"].isin(selected_periods)].copy()

        n_files = acc["Accounting Period"].nunique()

        # Reconciliation check
        raw_sum = st.session_state.tool2_raw_amount_sum
        export_sum = acc["Amount"].sum(min_count=1) if "Amount" in acc.columns else 0.0
        if abs(raw_sum - export_sum) < 0.01:
            st.success(f"Reconciliation passed — Amount totals match: {raw_sum:,.2f}")
        else:
            st.error(
                f"Reconciliation failed — Raw files total: {raw_sum:,.2f} | "
                f"Export total: {export_sum:,.2f} | "
                f"Difference: {raw_sum - export_sum:,.2f}"
            )

        # Apply Property Owner Type if mapping was provided
        include_owner_type = st.session_state.tool2_owner_type_map is not None
        if include_owner_type:
            owner_type_map = st.session_state.tool2_owner_type_map
            acc["Property Owner Type"] = acc["Owner"].map(lambda o: owner_type_map.get(o, "Third Party"))

        include_dept = st.session_state.tool2_group_by_dept and "Department" in acc.columns
        base_cols = ["Accounting Period", "Account", "Department", "Property", "Owner"] if include_dept else ["Accounting Period", "Account", "Property", "Owner"]
        if include_owner_type:
            base_cols = base_cols + ["Property Owner Type", "Amount"]
        else:
            base_cols = base_cols + ["Amount"]
        columns_to_keep = base_cols
        df_export = acc[[col for col in columns_to_keep if col in acc.columns]]

        df_export = df_export.copy()
        df_export["Owner"] = '="' + df_export["Owner"].astype(str).str.replace('"', '""') + '"'
        df_export["Property"] = '="' + df_export["Property"].astype(str).str.replace('"', '""') + '"'

        csv_data = df_export.to_csv(index=False, quoting=csv.QUOTE_MINIMAL).encode("utf-8")

        st.success(f"Ready to export — {len(df_export):,} rows across {n_files} file(s)")
        st.dataframe(df_export, use_container_width=True)

        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name="cleaned_data.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )

        st.divider()

        st.subheader("Audit — Raw Transactions")
        st.caption("All individual transactions before grouping, with Owner, Property, and Property Owner Type columns appended.")

        raw_audit = st.session_state.tool2_raw_transactions.copy()
        if include_owner_type and not raw_audit.empty:
            raw_audit["Property Owner Type"] = raw_audit["Owner"].map(lambda o: owner_type_map.get(o, "Third Party"))

        audit_cols_to_drop = [c for c in ["_group_by_dept", "Accounting Period"] if c in raw_audit.columns]
        raw_audit = raw_audit.drop(columns=audit_cols_to_drop, errors="ignore")

        st.dataframe(raw_audit, use_container_width=True)
        csv_audit = raw_audit.to_csv(index=False, quoting=csv.QUOTE_MINIMAL).encode("utf-8")
        st.download_button(
            label="Download Raw Transactions CSV",
            data=csv_audit,
            file_name="raw_transactions_audit.csv",
            mime="text/csv",
            use_container_width=True,
        )

        st.divider()

        if st.button("Restart", use_container_width=True):
            st.session_state.tool2_step = "upload"
            st.session_state.tool2_accumulated = pd.DataFrame()
            st.session_state.tool2_merges = []
            st.session_state.tool2_group_by_dept = False
            st.session_state.tool2_owner_type_map = None
            st.session_state.tool2_raw_amount_sum = 0.0
            st.session_state.tool2_raw_transactions = pd.DataFrame()
            st.session_state.pop("tool2_period_multiselect", None)
            st.rerun()


### ── TOOL 3: COMPANY EXPENSES DATA PREP ─────────────────────────────────────

elif st.session_state.tool == "tool3":

    import csv

    st.title("Company Expenses Data Prep")
    if st.button("← Back to Menu", key="back_tool3"):
        go_home()
        st.rerun()

    st.divider()

    # ── STEP: UPLOAD ──────────────────────────────────────────────────────────

    if st.session_state.tool3_step == "upload":

        if not st.session_state.tool3_accumulated.empty:
            n = st.session_state.tool3_accumulated["Accounting Period"].nunique()
            st.info(f"{n} file(s) already loaded. Upload the next file.")

        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"], key="tool3_uploader")

        if uploaded_file is not None:
            with st.spinner("Processing..."):
                df = pd.read_excel(uploaded_file, header=None)

                # Remove first 3 rows, use 4th row as header
                df = df.iloc[3:].reset_index(drop=True)
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)

                # Remove first 4 columns
                df = df.iloc[:, 4:]

                # Remove rows where Account is empty
                df.columns = df.columns.str.strip()
                df = df[df["Account"].notna()]
                df = df[df["Account"] != ""]
                df = df.reset_index(drop=True)

                # Create Accounting Period from Date
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
                df["Accounting Period"] = df["Date"] + pd.offsets.MonthEnd(0)
                df = df[df["Accounting Period"].notna()].reset_index(drop=True)

                # Create Owner and Property from Name / Class
                owners = []
                properties = []
                for _, row in df.iterrows():
                    name_value = str(row["Name"])
                    class_value = str(row["Class"])
                    if ":" in name_value:
                        parts = name_value.split(":", 1)
                        owners.append(parts[0].strip())
                        properties.append(parts[1].strip())
                    elif ":" in class_value:
                        parts = class_value.split(":", 1)
                        owners.append(parts[0].strip())
                        properties.append(parts[1].strip())
                    else:
                        owners.append(class_value.strip())
                        properties.append(class_value.strip())

                df["Owner"] = owners
                df["Property"] = properties

                # Clean Owner and Property
                df["Owner"] = df["Owner"].str.replace(" -C$", "", regex=True).str.strip().str.replace(r"\s+", " ", regex=True)
                df["Property"] = df["Property"].str.replace("XXX", "", regex=False).str.replace(r"\.\d{2}", "", regex=True).str.strip().str.replace(r"\s+", " ", regex=True)

                # Force Amount to numeric in case it was read as text
                if "Amount" in df.columns:
                    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")

                # Capture raw amount sum for reconciliation
                st.session_state.tool3_raw_amount_sum += df["Amount"].sum(min_count=1) if "Amount" in df.columns else 0.0

            st.session_state.tool3_accumulated = pd.concat(
                [st.session_state.tool3_accumulated, df], ignore_index=True
            )
            st.session_state.tool3_step = "action"
            st.rerun()

    # ── STEP: ACTION ──────────────────────────────────────────────────────────

    elif st.session_state.tool3_step == "action":

        acc = st.session_state.tool3_accumulated
        n_files = acc["Accounting Period"].nunique()

        st.success(f"{n_files} file(s) loaded — {len(acc):,} total rows")
        st.dataframe(acc, use_container_width=True)

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Add Another File", use_container_width=True, key="tool3_add_file"):
                st.session_state.tool3_step = "upload"
                st.rerun()
        with col2:
            if st.button("Continue →", type="primary", use_container_width=True, key="tool3_continue"):
                st.session_state.tool3_step = "account_mapping"
                st.rerun()

    # ── STEP: ACCOUNT MAPPING ─────────────────────────────────────────────────

    elif st.session_state.tool3_step == "account_mapping":

        st.subheader("Map Accounts to Expense Category")
        st.caption(
            "Review the default mapping below. Accounts not in the default list are shown prominently — "
            "assign them a category before continuing."
        )

        all_accounts = sorted(st.session_state.tool3_accumulated["Account"].dropna().unique().tolist())

        # Build working map from session state or defaults
        if st.session_state.tool3_account_map is None:
            working_map = {a: TOOL3_DEFAULT_ACCOUNT_MAP.get(a, "Uncategorized") for a in all_accounts}
        else:
            working_map = dict(st.session_state.tool3_account_map)
            for a in all_accounts:
                if a not in working_map:
                    working_map[a] = TOOL3_DEFAULT_ACCOUNT_MAP.get(a, "Uncategorized")

        unmapped = [a for a in all_accounts if working_map.get(a) == "Uncategorized"]
        mapped = [a for a in all_accounts if working_map.get(a) != "Uncategorized"]

        # Show unmapped accounts prominently
        if unmapped:
            st.error(
                f"**{len(unmapped)} account(s) are unmapped and must be assigned a category before you can continue:**\n\n"
                + "\n".join(f"- `{a}`" for a in unmapped)
            )
            st.divider()
            st.write("**Assign a category to an unmapped account:**")
            col_a, col_b = st.columns(2)
            with col_a:
                account_to_map = st.selectbox(
                    "Select unmapped account",
                    options=unmapped,
                    key="tool3_unmap_select",
                )
            with col_b:
                new_cat = st.selectbox(
                    "Assign category",
                    options=TOOL3_EXPENSE_CATEGORIES,
                    key="tool3_unmap_cat",
                )
            if st.button("Map Account", type="primary", key="tool3_map_account"):
                working_map[account_to_map] = new_cat
                st.session_state.tool3_account_map = working_map
                st.rerun()
            st.divider()
        else:
            st.success(f"All {len(all_accounts)} accounts are mapped.")

        # Show current mapping grouped by category
        if mapped:
            st.write("**Current mapping (accounts in your data):**")
            by_category = {}
            for a in mapped:
                cat = working_map[a]
                by_category.setdefault(cat, []).append(a)
            for cat in sorted(by_category):
                st.markdown(f"**{cat}**")
                for a in by_category[cat]:
                    st.markdown(f"- `{a}`")

        st.divider()
        st.write("**Change an existing mapping:**")
        col_a, col_b = st.columns(2)
        with col_a:
            account_to_change = st.selectbox(
                "Select account",
                options=all_accounts,
                key="tool3_change_account",
            )
        with col_b:
            current_cat = working_map.get(account_to_change, "Uncategorized")
            cat_options = TOOL3_EXPENSE_CATEGORIES
            default_idx = cat_options.index(current_cat) if current_cat in cat_options else 0
            new_cat_change = st.selectbox(
                "New category",
                options=cat_options,
                index=default_idx,
                key="tool3_change_cat",
            )
        if st.button("Update Mapping", key="tool3_update_mapping"):
            working_map[account_to_change] = new_cat_change
            st.session_state.tool3_account_map = working_map
            st.rerun()

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool3_acctmap_back"):
                st.session_state.tool3_step = "action"
                st.rerun()
        with col2:
            if unmapped:
                st.button(
                    "Apply & Continue →",
                    type="primary",
                    use_container_width=True,
                    key="tool3_acctmap_apply",
                    disabled=True,
                )
                st.caption("Map all accounts above before continuing.")
            else:
                if st.button("Apply & Continue →", type="primary", use_container_width=True, key="tool3_acctmap_apply"):
                    st.session_state.tool3_account_map = working_map
                    st.session_state.tool3_step = "dept_remap"
                    st.rerun()

    # ── STEP: DEPARTMENT REMAPPING ────────────────────────────────────────────

    elif st.session_state.tool3_step == "dept_remap":

        st.subheader("Rename Department Values")
        st.caption("Each department found in your data is shown below. Edit any name you want to rename — leave as-is to keep it unchanged.")

        acc = st.session_state.tool3_accumulated
        all_departments = sorted(acc["Department"].dropna().unique().tolist()) if "Department" in acc.columns else []

        if not all_departments:
            st.info("No Department values found in the data.")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("← Back", use_container_width=True, key="tool3_dept_back"):
                    st.session_state.tool3_step = "account_mapping"
                    st.rerun()
            with col2:
                if st.button("Continue →", type="primary", use_container_width=True, key="tool3_dept_apply"):
                    st.session_state.tool3_step = "owner_type"
                    st.rerun()
        else:
            if st.session_state.get("tool3_dept_remap") is None:
                working_remap = {d: d for d in all_departments}
            else:
                working_remap = dict(st.session_state.tool3_dept_remap)
                for d in all_departments:
                    if d not in working_remap:
                        working_remap[d] = d

            sorted_depts = sorted(working_remap.keys())

            for i, dept in enumerate(sorted_depts):
                key = f"tool3_dept_input_{i}"
                if key not in st.session_state:
                    st.session_state[key] = working_remap[dept]

            st.write(f"**{len(sorted_depts)} department(s) found:**")
            cols = st.columns(2)
            for i, dept in enumerate(sorted_depts):
                with cols[i % 2]:
                    st.text_input(f"`{dept}`", key=f"tool3_dept_input_{i}")

            st.divider()
            col1, col2 = st.columns(2)
            with col1:
                if st.button("← Back", use_container_width=True, key="tool3_dept_back"):
                    st.session_state.tool3_step = "account_mapping"
                    st.rerun()
            with col2:
                if st.button("Apply & Continue →", type="primary", use_container_width=True, key="tool3_dept_apply"):
                    new_remap = {}
                    for i, dept in enumerate(sorted_depts):
                        new_name = st.session_state.get(f"tool3_dept_input_{i}", dept).strip() or dept
                        new_remap[dept] = new_name
                    st.session_state.tool3_dept_remap = new_remap
                    st.session_state.tool3_step = "owner_type"
                    st.rerun()

    # ── STEP: OWNER TYPE MAPPING ──────────────────────────────────────────────

    elif st.session_state.tool3_step == "owner_type":

        OWNED_DEFAULT = {"CBTS LP", "CIF LP", "PBC LP", "KES LP", "CBC LP", "CinCB LP", "SinCB LP"}
        MGMT_DEFAULT = {"All FL Units"}

        st.subheader("Map Property Owner Type")
        st.caption(
            "Assign each Owner to a category. Everything not listed below is mapped as **Third Party**."
        )

        all_owners = sorted(st.session_state.tool3_accumulated["Owner"].dropna().unique().tolist())

        if st.session_state.tool3_owner_type_map is None:
            working_map = {}
            for o in all_owners:
                if o in OWNED_DEFAULT:
                    working_map[o] = "Owned"
                elif o in MGMT_DEFAULT:
                    working_map[o] = "SICB Management"
                else:
                    working_map[o] = "Third Party"
        else:
            working_map = dict(st.session_state.tool3_owner_type_map)
            for o in all_owners:
                if o not in working_map:
                    working_map[o] = "Third Party"

        st.write("**Current mapping** (only Owned and SICB Management shown — all others are Third Party):")

        owned_owners = [o for o, t in working_map.items() if t == "Owned"]
        mgmt_owners = [o for o, t in working_map.items() if t == "SICB Management"]

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Owned**")
            for o in owned_owners:
                st.markdown(f"- `{o}`")
        with col2:
            st.markdown("**SICB Management**")
            for o in mgmt_owners:
                st.markdown(f"- `{o}`")

        st.divider()
        st.write("**Add or change a mapping:**")

        third_party_owners = [o for o in all_owners if working_map.get(o) == "Third Party"]

        if third_party_owners:
            col_a, col_b = st.columns(2)
            with col_a:
                owner_to_add = st.selectbox(
                    "Select an Owner (currently Third Party)",
                    options=third_party_owners,
                    key="tool3_owner_type_select",
                )
            with col_b:
                new_type = st.radio(
                    "Map to",
                    options=["Owned", "SICB Management"],
                    key="tool3_owner_type_radio",
                    horizontal=True,
                )
            if st.button("Add Mapping", type="primary", key="tool3_add_mapping"):
                working_map[owner_to_add] = new_type
                st.session_state.tool3_owner_type_map = working_map
                st.rerun()
        else:
            st.info("All owners are already mapped.")

        st.divider()
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool3_owner_back"):
                st.session_state.tool3_step = "dept_remap"
                st.rerun()
        with col2:
            if st.button("Skip (no owner type column)", use_container_width=True, key="tool3_owner_skip"):
                st.session_state.tool3_owner_type_map = None
                st.session_state.tool3_step = "export"
                st.rerun()
        with col3:
            if st.button("Apply & Continue →", type="primary", use_container_width=True, key="tool3_owner_apply"):
                st.session_state.tool3_owner_type_map = working_map
                st.session_state.tool3_step = "export"
                st.rerun()

    # ── STEP: EXPORT ──────────────────────────────────────────────────────────

    elif st.session_state.tool3_step == "export":

        acc = st.session_state.tool3_accumulated.copy()
        n_files = acc["Accounting Period"].nunique()

        # Reconciliation check
        raw_sum = st.session_state.tool3_raw_amount_sum
        export_sum = acc["Amount"].sum(min_count=1) if "Amount" in acc.columns else 0.0
        if abs(raw_sum - export_sum) < 0.01:
            st.success(f"Reconciliation passed — Amount totals match: {raw_sum:,.2f}")
        else:
            st.error(
                f"Reconciliation failed — Raw files total: {raw_sum:,.2f} | "
                f"Export total: {export_sum:,.2f} | "
                f"Difference: {raw_sum - export_sum:,.2f}"
            )

        # Apply Expense Category from account mapping
        account_map = st.session_state.tool3_account_map or {}
        acc["Expense Category"] = acc["Account"].map(lambda a: account_map.get(a, "Uncategorized"))

        # Apply Property Owner Type if mapping was provided
        include_owner_type = st.session_state.tool3_owner_type_map is not None
        if include_owner_type:
            owner_type_map = st.session_state.tool3_owner_type_map
            acc["Property Owner Type"] = acc["Owner"].map(lambda o: owner_type_map.get(o, "Third Party"))

        # Apply Department remap if provided
        if st.session_state.get("tool3_dept_remap") and "Department" in acc.columns:
            dept_remap = st.session_state.tool3_dept_remap
            acc["Department"] = acc["Department"].map(
                lambda d: dept_remap.get(str(d).strip(), str(d)) if pd.notna(d) else d
            )

        # ── FILTER BY ACCOUNTING PERIOD ───────────────────────────────────────
        all_periods = sorted(acc["Accounting Period"].dropna().unique().tolist(), key=lambda x: str(x))

        st.subheader("Filter by Accounting Period")
        st.caption("All periods are selected by default — deselect any you want to exclude from the export.")

        if "tool3_period_multiselect" not in st.session_state:
            st.session_state["tool3_period_multiselect"] = all_periods
        else:
            valid_p = [p for p in st.session_state["tool3_period_multiselect"] if p in all_periods]
            if set(valid_p) != set(st.session_state["tool3_period_multiselect"]):
                st.session_state["tool3_period_multiselect"] = all_periods

        col_sel_p, col_desel_p = st.columns(2)
        with col_sel_p:
            if st.button("Select All", key="tool3_period_select_all", use_container_width=True):
                st.session_state["tool3_period_multiselect"] = all_periods
                st.rerun()
        with col_desel_p:
            if st.button("Deselect All", key="tool3_period_desel_all", use_container_width=True):
                st.session_state["tool3_period_multiselect"] = []
                st.rerun()

        selected_periods = st.multiselect(
            "Accounting Periods to include in export",
            options=all_periods,
            key="tool3_period_multiselect",
            format_func=lambda x: str(x),
        )

        acc = acc[acc["Accounting Period"].isin(selected_periods)].copy()

        st.divider()

        # ── FILTER BY EXPENSE CATEGORY ────────────────────────────────────────
        all_expense_categories = sorted(acc["Expense Category"].dropna().unique().tolist())

        st.subheader("Filter by Expense Category")
        st.caption("All categories are selected by default — deselect any you want to exclude from the export.")

        if "tool3_cat_multiselect" not in st.session_state:
            st.session_state["tool3_cat_multiselect"] = all_expense_categories
        else:
            valid = [c for c in st.session_state["tool3_cat_multiselect"] if c in all_expense_categories]
            if set(valid) != set(st.session_state["tool3_cat_multiselect"]):
                st.session_state["tool3_cat_multiselect"] = all_expense_categories

        col_sel, col_desel = st.columns(2)
        with col_sel:
            if st.button("Select All", key="tool3_cat_select_all", use_container_width=True):
                st.session_state["tool3_cat_multiselect"] = all_expense_categories
                st.rerun()
        with col_desel:
            if st.button("Deselect All", key="tool3_cat_desel_all", use_container_width=True):
                st.session_state["tool3_cat_multiselect"] = []
                st.rerun()

        selected_categories = st.multiselect(
            "Expense Categories to include in export",
            options=all_expense_categories,
            key="tool3_cat_multiselect",
        )

        acc = acc[acc["Expense Category"].isin(selected_categories)].copy()

        st.divider()

        columns_to_keep = ["Accounting Period", "Account", "Department", "Property", "Owner"]
        if include_owner_type:
            columns_to_keep.append("Property Owner Type")
        columns_to_keep += ["Expense Category", "Memo", "Source Name", "Amount"]

        df_export = acc[[col for col in columns_to_keep if col in acc.columns]].copy()

        df_export["Owner"] = '="' + df_export["Owner"].astype(str).str.replace('"', '""') + '"'
        df_export["Property"] = '="' + df_export["Property"].astype(str).str.replace('"', '""') + '"'

        csv_data = df_export.to_csv(index=False, quoting=csv.QUOTE_MINIMAL).encode("utf-8")

        if not selected_periods or not selected_categories:
            st.warning("No periods or categories selected — select at least one of each above to export.")
        else:
            n_sel_p = len(selected_periods)
            n_tot_p = len(all_periods)
            n_sel_c = len(selected_categories)
            n_tot_c = len(all_expense_categories)
            period_label = f"all {n_tot_p} periods" if n_sel_p == n_tot_p else f"{n_sel_p} of {n_tot_p} periods"
            category_label = f"all {n_tot_c} categories" if n_sel_c == n_tot_c else f"{n_sel_c} of {n_tot_c} categories"
            st.success(f"Ready to export — {len(df_export):,} rows · {period_label} · {category_label}")
            st.dataframe(df_export, use_container_width=True)

            st.download_button(
                label="Download CSV",
                data=csv_data,
                file_name="company_expenses.csv",
                mime="text/csv",
                type="primary",
                use_container_width=True,
            )

        st.divider()

        if st.button("Restart", use_container_width=True, key="tool3_restart"):
            st.session_state.tool3_step = "upload"
            st.session_state.tool3_accumulated = pd.DataFrame()
            st.session_state.tool3_merges = []
            st.session_state.tool3_group_by_dept = False
            st.session_state.tool3_owner_type_map = None
            st.session_state.tool3_raw_amount_sum = 0.0
            st.session_state.tool3_account_map = None
            st.session_state.tool3_dept_remap = None
            for k in [k for k in st.session_state if k.startswith("tool3_dept_input_")]:
                del st.session_state[k]
            st.session_state.pop("tool3_period_multiselect", None)
            st.session_state.pop("tool3_cat_multiselect", None)
            st.rerun()


### ── TOOL 4: SICB MANAGEMENT EXPENSE ALLOCATION ─────────────────────────────

elif st.session_state.tool == "tool4":

    import csv

    st.title("SICB Management Expense Allocation")
    if st.button("← Back to Menu", key="back_tool4"):
        go_home()
        st.rerun()

    st.divider()

    # ── STEP: UPLOAD ──────────────────────────────────────────────────────────

    if st.session_state.tool4_step == "upload":

        st.write("Upload both files to proceed.")
        expenses_file = st.file_uploader("Income & Expenses CSV", type=["csv"], key="tool4_expenses_uploader")
        units_file = st.file_uploader("Property / Units CSV", type=["csv"], key="tool4_units_uploader")

        if expenses_file is not None and units_file is not None:
            with st.spinner("Loading files..."):
                expenses_df = pd.read_csv(expenses_file)
                units_df = pd.read_csv(units_file)

                # Capture dropped rows before filtering
                full_expenses_df = expenses_df.copy()
                full_expenses_df["Accounting Period"] = pd.to_datetime(full_expenses_df["Accounting Period"], errors="coerce")
                full_expenses_df = full_expenses_df[full_expenses_df["Accounting Period"].notna()].reset_index(drop=True)
                full_expenses_df["Accounting Period"] = full_expenses_df["Accounting Period"] + pd.offsets.MonthEnd(0)
                dropped_df = full_expenses_df[~full_expenses_df["Property Owner Type"].isin(["SICB Management", "All FL Units"])].copy()
                st.session_state.tool4_dropped_df = dropped_df

                # Filter to SICB Management and All FL Units
                expenses_df = full_expenses_df[full_expenses_df["Property Owner Type"].isin(["SICB Management", "All FL Units"])].copy().reset_index(drop=True)

                # Clean QuickBooks Name decimal suffixes (e.g. "BRR 7.00" → "BRR 7")
                if "QuickBooks Name" in units_df.columns:
                    units_df["QuickBooks Name"] = units_df["QuickBooks Name"].astype(str).str.replace(r"\.\d+$", "", regex=True).str.strip()

                # Parse unit dates
                units_df["Purchase/Onboarded Date"] = pd.to_datetime(units_df["Purchase/Onboarded Date"], errors="coerce")
                units_df["Offboarded Date"] = pd.to_datetime(units_df["Offboarded Date"], errors="coerce")

                st.session_state.tool4_expenses_df = expenses_df
                st.session_state.tool4_units_df = units_df

            st.session_state.tool4_step = "unit_check"
            st.rerun()

    # ── STEP: UNIT OWNER CHECK ────────────────────────────────────────────────

    elif st.session_state.tool4_step == "unit_check":

        OWNER_TYPE_OPTIONS = ["Owned", "Third Party", "SICB Management"]

        units_df = st.session_state.tool4_units_df
        overrides = dict(st.session_state.tool4_unit_owner_overrides)

        def _unit_key(row):
            qb = str(row.get("QuickBooks Name", "")).strip()
            if qb and qb.lower() != "nan":
                return qb
            un = str(row.get("Unit Name", "")).strip()
            if un and un.lower() != "nan":
                return un
            return f"row_{row.name}"

        def _effective_type(row):
            return overrides.get(_unit_key(row), {}).get("Owner Type") or str(row.get("Owner Type", "")).strip()

        missing_mask = units_df.apply(
            lambda r: not _effective_type(r) or _effective_type(r).lower() == "nan", axis=1
        )
        missing_units = units_df[missing_mask]

        if missing_units.empty:
            st.success(f"All {len(units_df)} units have Owner Type assigned — no manual mapping needed.")
        else:
            st.warning(
                f"{len(missing_units)} unit(s) are missing Owner Type. "
                "These will be excluded from **Owned Only** and **Third Party Only** allocation pools unless assigned below."
            )
            display_cols = [c for c in ["QuickBooks Name", "Unit Name", "Owner Name", "Owner Type", "Status"] if c in missing_units.columns]
            st.dataframe(missing_units[display_cols].reset_index(drop=True), use_container_width=True)

            st.divider()
            st.write("**Assign Owner Type:**")

            unit_labels = []
            unit_keys = []
            for _, row in missing_units.iterrows():
                unit_labels.append(_unit_key(row))
                unit_keys.append(_unit_key(row))

            col_a, col_b = st.columns(2)
            with col_a:
                selected_label = st.selectbox(
                    "Select unit",
                    options=unit_labels,
                    key="tool4_unit_check_select",
                )
            with col_b:
                selected_type = st.selectbox(
                    "Owner Type",
                    options=OWNER_TYPE_OPTIONS,
                    key="tool4_unit_check_type",
                )

            if st.button("Assign", type="primary", key="tool4_unit_check_assign"):
                overrides[selected_label] = {"Owner Type": selected_type}
                st.session_state.tool4_unit_owner_overrides = overrides
                st.rerun()

        if overrides:
            st.divider()
            st.write("**Current manual assignments:**")
            for key, vals in overrides.items():
                st.markdown(f"- `{key}` → **{vals.get('Owner Type', '')}**")
            if st.button("Clear All Assignments", key="tool4_unit_check_clear"):
                st.session_state.tool4_unit_owner_overrides = {}
                st.rerun()

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool4_unit_check_back"):
                st.session_state.tool4_step = "upload"
                st.rerun()
        with col2:
            continue_label = "Continue →" if missing_units.empty else "Continue anyway →"
            if st.button(continue_label, type="primary", use_container_width=True, key="tool4_unit_check_continue"):
                if overrides:
                    updated_units = st.session_state.tool4_units_df.copy()
                    for _, row in updated_units.iterrows():
                        key = _unit_key(row)
                        if key in overrides:
                            updated_units.at[row.name, "Owner Type"] = overrides[key]["Owner Type"]
                    st.session_state.tool4_units_df = updated_units
                st.session_state.tool4_step = "action"
                st.rerun()

    # ── STEP: ACTION ──────────────────────────────────────────────────────────

    elif st.session_state.tool4_step == "action":

        expenses_df = st.session_state.tool4_expenses_df
        units_df = st.session_state.tool4_units_df

        n_rows = len(expenses_df)
        n_months = expenses_df["Accounting Period"].nunique()
        st.success(f"{n_rows:,} expense rows loaded across {n_months} month(s)")

        st.subheader("Filtered Expenses Preview")
        st.dataframe(expenses_df, use_container_width=True)

        st.subheader("Active Unit Count by Month")
        st.caption("A unit is active if Purchase/Onboarded Date ≤ last day of month and Offboarded Date is null or after the last day of month.")
        months = sorted(expenses_df["Accounting Period"].dropna().unique())
        owned_entities = sorted(
            units_df.loc[units_df["Owner Type"] == "Owned", "Owner"].dropna().unique().tolist()
        )
        unit_count_rows = []
        for month in months:
            active = units_df[
                (units_df["Purchase/Onboarded Date"].notna()) &
                (units_df["Purchase/Onboarded Date"] <= month) &
                (units_df["Offboarded Date"].isna() | (units_df["Offboarded Date"] > month))
            ]
            row = {"Accounting Period": month.date()}
            row["Total"] = len(active)
            row["Third Party"] = int((active["Owner Type"] == "Third Party").sum())
            row["SICB Management"] = int((active["Owner Type"] == "SICB Management").sum())
            row["Owned"] = int((active["Owner Type"] == "Owned").sum())
            for entity in owned_entities:
                row[f"Owned - {entity}"] = int(
                    ((active["Owner Type"] == "Owned") & (active["Owner"] == entity)).sum()
                )
            unit_count_rows.append(row)
        unit_count_df = pd.DataFrame(unit_count_rows)
        st.dataframe(unit_count_df, use_container_width=True)

        if unit_count_df["Total"].eq(0).any():
            st.warning("Some months have 0 active units — those expense rows will be excluded from the output.")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Start Over", use_container_width=True, key="tool4_back_upload"):
                st.session_state.tool4_step = "upload"
                st.session_state.tool4_expenses_df = pd.DataFrame()
                st.session_state.tool4_units_df = pd.DataFrame()
                st.rerun()
        with col2:
            if st.button("Continue →", type="primary", use_container_width=True, key="tool4_continue"):
                st.session_state.tool4_step = "dept_accounts"
                st.rerun()

    # ── STEP: DEPT FILTER ACCOUNTS ────────────────────────────────────────────

    elif st.session_state.tool4_step == "dept_accounts":

        st.subheader("Department-Based Allocation")
        st.caption(
            "Transactions for the accounts below will be allocated to **Third Party units only** "
            "if their department matches one of the listed departments. "
            "All other departments for these accounts are allocated across all units."
        )

        dept_filter_accounts = list(st.session_state.tool4_dept_filter_accounts)
        dept_filter_departments = list(st.session_state.tool4_dept_filter_departments)

        st.write("**Departments → Third Party Only:**")
        for i, dept in enumerate(dept_filter_departments):
            col_d, col_r = st.columns([6, 1])
            with col_d:
                st.code(dept)
            with col_r:
                if st.button("Remove", key=f"tool4_remove_dept_{i}"):
                    dept_filter_departments.pop(i)
                    st.session_state.tool4_dept_filter_departments = dept_filter_departments
                    st.rerun()

        new_dept = st.text_input("Add a department", key="tool4_new_dept_input")
        if st.button("Add Department", key="tool4_add_dept"):
            nd = new_dept.strip()
            if nd and nd not in dept_filter_departments:
                dept_filter_departments.append(nd)
                st.session_state.tool4_dept_filter_departments = dept_filter_departments
                st.rerun()

        st.divider()

        st.write("**Accounts subject to department-based allocation:**")
        for i, acct in enumerate(dept_filter_accounts):
            col_a, col_r = st.columns([6, 1])
            with col_a:
                st.markdown(f"- `{acct}`")
            with col_r:
                if st.button("Remove", key=f"tool4_remove_acct_{i}"):
                    dept_filter_accounts.pop(i)
                    st.session_state.tool4_dept_filter_accounts = dept_filter_accounts
                    st.rerun()

        all_expense_accounts = sorted(st.session_state.tool4_expenses_df["Account"].dropna().unique().tolist())
        addable_accounts = [a for a in all_expense_accounts if a not in dept_filter_accounts]
        if addable_accounts:
            add_acct = st.selectbox("Add an account", options=[""] + addable_accounts, key="tool4_add_acct_select")
            if st.button("Add Account", key="tool4_add_acct_btn"):
                if add_acct and add_acct not in dept_filter_accounts:
                    dept_filter_accounts.append(add_acct)
                    st.session_state.tool4_dept_filter_accounts = dept_filter_accounts
                    st.rerun()

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool4_dept_back"):
                st.session_state.tool4_step = "action"
                st.rerun()
        with col2:
            if st.button("Continue →", type="primary", use_container_width=True, key="tool4_dept_continue"):
                st.session_state.tool4_dept_filter_accounts = dept_filter_accounts
                st.session_state.tool4_dept_filter_departments = dept_filter_departments
                st.session_state.tool4_step = "account_mode"
                st.rerun()

    # ── STEP: ACCOUNT ALLOCATION MODE ────────────────────────────────────────

    elif st.session_state.tool4_step == "account_mode":

        TOOL4_OVERRIDE_MODES = ["Owned Only", "Third Party Only", "Keep as SICB Management"]

        st.subheader("Account Allocation Mode")
        st.caption(
            "All remaining accounts are split across **All Units** by default. "
            "Override specific accounts below."
        )

        expenses_df = st.session_state.tool4_expenses_df
        dept_filter_accounts = st.session_state.tool4_dept_filter_accounts
        all_accounts = sorted(expenses_df["Account"].dropna().unique().tolist())
        remaining_accounts = [a for a in all_accounts if a not in dept_filter_accounts]

        working_map = {k: v for k, v in st.session_state.tool4_account_mode_map.items() if k in remaining_accounts}

        owned_only = [a for a in remaining_accounts if working_map.get(a) == "Owned Only"]
        tp_only = [a for a in remaining_accounts if working_map.get(a) == "Third Party Only"]
        sicb_only = [a for a in remaining_accounts if working_map.get(a) == "Keep as SICB Management"]

        if owned_only or tp_only or sicb_only:
            st.write("**Current overrides:**")
            for label, lst in [("Owned Only", owned_only), ("Third Party Only", tp_only), ("Keep as SICB Management", sicb_only)]:
                if lst:
                    st.markdown(f"*{label}*")
                    for a in lst:
                        col_a, col_r = st.columns([6, 1])
                        with col_a:
                            st.markdown(f"- `{a}`")
                        with col_r:
                            if st.button("Remove", key=f"tool4_rm_override_{a}"):
                                working_map.pop(a, None)
                                st.session_state.tool4_account_mode_map = working_map
                                st.rerun()
            st.divider()

        overridable = [a for a in remaining_accounts if a not in working_map]
        if overridable:
            st.write("**Add an override:**")
            col_a, col_b = st.columns(2)
            with col_a:
                account_to_override = st.selectbox("Account", options=overridable, key="tool4_override_account")
            with col_b:
                override_mode = st.selectbox("Mode", options=TOOL4_OVERRIDE_MODES, key="tool4_override_mode")
            if st.button("Add Override", type="primary", key="tool4_add_override"):
                working_map[account_to_override] = override_mode
                st.session_state.tool4_account_mode_map = working_map
                st.rerun()
        else:
            st.info("All accounts have overrides.")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool4_mode_back"):
                st.session_state.tool4_step = "dept_accounts"
                st.rerun()
        with col2:
            if st.button("Allocate & Continue →", type="primary", use_container_width=True, key="tool4_mode_apply"):
                st.session_state.tool4_account_mode_map = working_map
                units_df = st.session_state.tool4_units_df
                dept_filter_accts = st.session_state.tool4_dept_filter_accounts
                dept_filter_depts = st.session_state.tool4_dept_filter_departments
                months = sorted(expenses_df["Accounting Period"].dropna().unique())

                with st.spinner("Allocating expenses..."):
                    output_chunks = []
                    for month in months:
                        cols_needed = [c for c in ["Account", "Department", "Amount"] if c in expenses_df.columns]
                        month_expenses = expenses_df[expenses_df["Accounting Period"] == month][cols_needed].copy()
                        all_active = units_df[
                            (units_df["Purchase/Onboarded Date"].notna()) &
                            (units_df["Purchase/Onboarded Date"] <= month) &
                            (units_df["Offboarded Date"].isna() | (units_df["Offboarded Date"] > month))
                        ][["QuickBooks Name", "Owner", "Owner Type"]].copy()

                        for _, exp_row in month_expenses.iterrows():
                            account = exp_row["Account"]
                            amount = exp_row["Amount"]
                            _dept = exp_row.get("Department", "")
                            department = "" if pd.isna(_dept) else str(_dept).strip()

                            if account in dept_filter_accts and department in dept_filter_depts:
                                pool = all_active[all_active["Owner Type"] == "Third Party"].copy()
                            else:
                                mode = working_map.get(account, "All Units")
                                if mode == "Keep as SICB Management":
                                    output_chunks.append(pd.DataFrame([{
                                        "Accounting Period": month.date(),
                                        "Account": account,
                                        "Department": department,
                                        "Property": "SICB Management",
                                        "Allocated Amount": amount,
                                        "Owner": "",
                                        "Property Owner Type": "SICB Management",
                                    }]))
                                    continue
                                elif mode == "Owned Only":
                                    pool = all_active[all_active["Owner Type"] == "Owned"].copy()
                                elif mode == "Third Party Only":
                                    pool = all_active[all_active["Owner Type"] == "Third Party"].copy()
                                else:
                                    pool = all_active.copy()

                            n_units = len(pool)
                            if n_units == 0:
                                continue

                            crossed = pd.DataFrame([{"Account": account, "Department": department, "Amount": amount}]).merge(
                                pool[["QuickBooks Name", "Owner", "Owner Type"]], how="cross"
                            )
                            unit_amount = round(amount / n_units, 2)
                            allocated = [unit_amount] * n_units
                            allocated[-1] = round(amount - unit_amount * (n_units - 1), 2)
                            crossed["Allocated Amount"] = allocated
                            crossed["Accounting Period"] = month.date()
                            crossed = crossed.rename(columns={"QuickBooks Name": "Property", "Owner Type": "Property Owner Type"})
                            output_chunks.append(
                                crossed[["Accounting Period", "Account", "Department", "Property", "Allocated Amount", "Owner", "Property Owner Type"]]
                            )

                    if output_chunks:
                        result_df = pd.concat(output_chunks, ignore_index=True)
                    else:
                        result_df = pd.DataFrame(columns=["Accounting Period", "Account", "Department", "Property", "Allocated Amount", "Owner", "Property Owner Type"])

                st.session_state.tool4_result_df = result_df
                st.session_state.tool4_missing_overrides = {}
                st.session_state.tool4_step = "missing_check"
                st.rerun()

    # ── STEP: MISSING OWNER CHECK ─────────────────────────────────────────────

    elif st.session_state.tool4_step == "missing_check":

        result_df = st.session_state.tool4_result_df.copy()
        overrides = dict(st.session_state.tool4_missing_overrides)

        OWNER_TYPE_OPTIONS = ["Owned", "Third Party", "SICB Management"]

        def _apply_overrides(df, ovr):
            df = df.copy()
            for prop, vals in ovr.items():
                mask = df["Property"] == prop
                if "Owner" in vals:
                    df.loc[mask, "Owner"] = vals["Owner"]
                if "Property Owner Type" in vals:
                    df.loc[mask, "Property Owner Type"] = vals["Property Owner Type"]
            return df

        result_df = _apply_overrides(result_df, overrides)

        missing_mask = (
            result_df["Owner"].isna() | (result_df["Owner"].astype(str).str.strip() == "") |
            result_df["Property Owner Type"].isna() | (result_df["Property Owner Type"].astype(str).str.strip() == "")
        )
        missing_props = sorted(result_df.loc[missing_mask, "Property"].dropna().unique().tolist())

        if not missing_props:
            st.success("No missing Owner or Property Owner Type — all units are complete.")
        else:
            st.warning(f"{len(missing_props)} unit(s) are missing Owner or Property Owner Type:")
            for p in missing_props:
                st.markdown(f"- `{p}`")

            st.divider()
            st.write("**Assign values:**")
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                selected_prop = st.selectbox("Unit (QuickBooks Name)", options=missing_props, key="tool4_missing_prop")
            with col_b:
                all_owners = sorted(st.session_state.tool4_units_df["Owner"].dropna().unique().tolist())
                selected_owner = st.selectbox("Owner", options=[""] + all_owners, key="tool4_missing_owner")
            with col_c:
                selected_ot = st.selectbox("Property Owner Type", options=[""] + OWNER_TYPE_OPTIONS, key="tool4_missing_ot")

            if st.button("Assign", type="primary", key="tool4_missing_assign"):
                entry = overrides.get(selected_prop, {})
                if selected_owner:
                    entry["Owner"] = selected_owner
                if selected_ot:
                    entry["Property Owner Type"] = selected_ot
                overrides[selected_prop] = entry
                st.session_state.tool4_missing_overrides = overrides
                st.rerun()

        if overrides:
            st.divider()
            st.write("**Current assignments:**")
            for prop, vals in overrides.items():
                st.markdown(f"- `{prop}` → Owner: **{vals.get('Owner', '—')}** | Type: **{vals.get('Property Owner Type', '—')}**")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back", use_container_width=True, key="tool4_missing_back"):
                st.session_state.tool4_step = "account_mode"
                st.rerun()
        with col2:
            label = "Continue →" if missing_props else "Continue →"
            if st.button(label, type="primary", use_container_width=True, key="tool4_missing_continue"):
                st.session_state.tool4_result_df = _apply_overrides(st.session_state.tool4_result_df, overrides)
                st.session_state.tool4_step = "export"
                st.rerun()

    # ── STEP: EXPORT ──────────────────────────────────────────────────────────

    elif st.session_state.tool4_step == "export":

        result_df = st.session_state.tool4_result_df

        original_sum = st.session_state.tool4_expenses_df["Amount"].sum(min_count=1)
        allocated_sum = result_df["Allocated Amount"].sum(min_count=1) if not result_df.empty else 0.0
        if abs(original_sum - allocated_sum) < 0.01:
            st.success(f"Reconciliation passed — totals match: {original_sum:,.2f}")
        else:
            st.error(
                f"Reconciliation failed — Original: {original_sum:,.2f} | "
                f"Allocated: {allocated_sum:,.2f} | "
                f"Difference: {original_sum - allocated_sum:,.2f}"
            )

        if result_df.empty:
            st.warning("No output rows generated — check that your unit file has active units for the months in your expense file.")
        else:
            include_dept = st.checkbox("Include Department in grouping", value=False, key="tool4_export_include_dept")
            group_keys = ["Accounting Period", "Account", "Department", "Property"] if include_dept else ["Accounting Period", "Account", "Property"]
            df_export = (
                result_df.groupby(group_keys, as_index=False)
                .agg({"Allocated Amount": "sum", "Owner": "first", "Property Owner Type": "first"})
            )
            col_order = group_keys + ["Allocated Amount", "Owner", "Property Owner Type"]
            df_export = df_export[col_order]

            st.success(f"Ready to export — {len(df_export):,} rows")
            st.dataframe(df_export, use_container_width=True)

            df_export_csv = df_export.copy()
            df_export_csv["Property"] = '="' + df_export_csv["Property"].astype(str).str.replace('"', '""') + '"'
            csv_data = df_export_csv.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="Download CSV",
                data=csv_data,
                file_name="sicb_expense_allocation.csv",
                mime="text/csv",
                type="primary",
                use_container_width=True,
            )

        st.divider()

        st.subheader("Combined Export")
        st.caption(
            "Allocated rows combined with the original non-SICB/All FL Units transactions, "
            "grouped by Accounting Period + Account + Property, Department dropped."
        )

        dropped_df = st.session_state.tool4_dropped_df
        if not result_df.empty or not dropped_df.empty:
            allocated_part = result_df.rename(columns={"Allocated Amount": "Amount"}).copy()
            allocated_part["Accounting Period"] = pd.to_datetime(allocated_part["Accounting Period"]).dt.to_period("M").dt.to_timestamp("M")
            allocated_part = allocated_part[["Accounting Period", "Account", "Property", "Amount", "Owner", "Property Owner Type"]]

            if not dropped_df.empty:
                dropped_part = dropped_df.copy()
                dropped_part["Amount"] = pd.to_numeric(dropped_part["Amount"], errors="coerce")
                dropped_part["Accounting Period"] = pd.to_datetime(dropped_part["Accounting Period"], errors="coerce") + pd.offsets.MonthEnd(0)
                dropped_part["Accounting Period"] = dropped_part["Accounting Period"].dt.to_period("M").dt.to_timestamp("M")
                dropped_part["Property"] = dropped_part["Property"].astype(str).str.replace(r'^="(.*)"$', r'\1', regex=True).str.replace('""', '"')
                dropped_part["Owner"] = dropped_part["Owner"].astype(str).str.replace(r'^="(.*)"$', r'\1', regex=True).str.replace('""', '"')
                dropped_part = dropped_part[["Accounting Period", "Account", "Property", "Amount", "Owner", "Property Owner Type"]]
            else:
                dropped_part = pd.DataFrame(columns=["Accounting Period", "Account", "Property", "Amount", "Owner", "Property Owner Type"])

            combined = pd.concat([allocated_part, dropped_part], ignore_index=True)
            combined = (
                combined.groupby(["Accounting Period", "Account", "Property"], as_index=False)
                .agg({"Amount": "sum", "Owner": "first", "Property Owner Type": "first"})
            )

            st.success(f"{len(combined):,} rows")
            st.dataframe(combined, use_container_width=True)

            combined_csv = combined.copy()
            combined_csv["Property"] = '="' + combined_csv["Property"].astype(str).str.replace('"', '""') + '"'
            csv_combined = combined_csv.to_csv(index=False).encode("utf-8")
            st.download_button(
                label="Download Combined CSV",
                data=csv_combined,
                file_name="combined_expense_allocation.csv",
                mime="text/csv",
                type="primary",
                use_container_width=True,
            )

        st.divider()

        if st.button("Restart", use_container_width=True, key="tool4_restart"):
            st.session_state.tool4_step = "upload"
            st.session_state.tool4_expenses_df = pd.DataFrame()
            st.session_state.tool4_units_df = pd.DataFrame()
            st.session_state.tool4_result_df = pd.DataFrame()
            st.session_state.tool4_account_mode_map = {}
            st.session_state.tool4_unit_owner_overrides = {}
            st.session_state.tool4_dept_filter_accounts = list(TOOL4_DEFAULT_DEPT_FILTER_ACCOUNTS)
            st.session_state.tool4_dept_filter_departments = list(TOOL4_DEFAULT_DEPT_FILTER_DEPARTMENTS)
            st.session_state.tool4_missing_overrides = {}
            st.session_state.tool4_dropped_df = pd.DataFrame()
            st.rerun()


### ── TOOL 5: FINAL PROPERTY LEVEL P&L ────────────────────────────────────────

elif st.session_state.tool == "tool5":

    st.title("Final Property Level P&L")
    if st.button("← Back to Menu", key="back_tool5"):
        go_home()
        st.rerun()

    st.divider()

    # ── STEP: UPLOAD ──────────────────────────────────────────────────────────

    if st.session_state.tool5_step == "upload":

        st.subheader("Step 1 — 3-Way Reconciliation")
        st.caption(
            "Upload the Portfolio-level P&L, Property-level P&L, and GL transaction detail for each period "
            "(e.g. one period per year, up to 3). Period 1 is required; Periods 2 and 3 are optional. Each "
            "period's 3 files are reconciled against each other independently, then combined for the unit "
            "economics extraction and export."
        )

        period_uploads = []
        for i in range(1, 4):
            label = "required" if i == 1 else "optional"
            st.write(f"**Period {i}** ({label})")
            col1, col2, col3 = st.columns(3)
            with col1:
                portfolio_file = st.file_uploader("Portfolio-Level P&L (Excel)", type=["xlsx", "xls"], key=f"tool5_portfolio_uploader_{i}")
            with col2:
                property_file = st.file_uploader("Property-Level P&L (Excel)", type=["xlsx", "xls"], key=f"tool5_property_uploader_{i}")
            with col3:
                gl_file = st.file_uploader("GL Transaction Detail (Excel)", type=["xlsx", "xls"], key=f"tool5_gl_uploader_{i}")
            period_uploads.append((portfolio_file, property_file, gl_file))
            st.divider()

        validation_errors = []
        populated_periods = []
        for i, (portfolio_file, property_file, gl_file) in enumerate(period_uploads, start=1):
            n_present = sum(f is not None for f in (portfolio_file, property_file, gl_file))
            if n_present == 0:
                continue
            if n_present < 3:
                validation_errors.append(f"Period {i} has only {n_present} of 3 files uploaded — upload all 3, or none, for this period.")
            else:
                populated_periods.append((i, portfolio_file, property_file, gl_file))

        for msg in validation_errors:
            st.warning(msg)

        period1_ready = period_uploads[0][0] is not None and period_uploads[0][1] is not None and period_uploads[0][2] is not None

        if period1_ready and not validation_errors:
            if st.button("Run Reconciliation", type="primary", use_container_width=True):
                portfolio_raw_list, property_raw_list, gl_raw_list = [], [], []
                recon_dfs, extra_gl_dfs = [], []
                had_error = False
                with st.spinner("Reconciling..."):
                    for period_num, portfolio_file, property_file, gl_file in populated_periods:
                        try:
                            portfolio_raw = pd.read_excel(portfolio_file, header=None)
                            property_raw = pd.read_excel(property_file, header=None)
                            gl_raw = pd.read_excel(gl_file, header=0)
                            recon_df, extra_gl_df = tool5_reconcile(portfolio_raw, property_raw, gl_raw)
                        except ValueError as e:
                            st.error(f"Period {period_num} error: {e}")
                            had_error = True
                            break
                        recon_df = recon_df.copy()
                        recon_df.insert(0, "Period", f"Period {period_num}")
                        extra_gl_df = extra_gl_df.copy()
                        extra_gl_df.insert(0, "Period", f"Period {period_num}")
                        portfolio_raw_list.append(portfolio_raw)
                        property_raw_list.append(property_raw)
                        gl_raw_list.append(gl_raw)
                        recon_dfs.append(recon_df)
                        extra_gl_dfs.append(extra_gl_df)

                if not had_error:
                    st.session_state.tool5_recon_df = pd.concat(recon_dfs, ignore_index=True)
                    st.session_state.tool5_extra_gl_df = pd.concat(extra_gl_dfs, ignore_index=True)
                    st.session_state.tool5_portfolio_raw_list = portfolio_raw_list
                    st.session_state.tool5_property_raw_list = property_raw_list
                    st.session_state.tool5_gl_raw_list = gl_raw_list
                    st.session_state.tool5_step = "reconcile"
                    st.rerun()

    # ── STEP: RECONCILE ───────────────────────────────────────────────────────

    elif st.session_state.tool5_step == "reconcile":

        recon_df = st.session_state.tool5_recon_df
        extra_gl_df = st.session_state.tool5_extra_gl_df

        n_total = len(recon_df)
        n_match = int((recon_df["Checker"] == "Match").sum())
        n_expected = int(recon_df["Checker"].str.contains("expected", na=False).sum())
        n_real_mismatch = n_total - n_match - n_expected

        col1, col2, col3 = st.columns(3)
        col1.metric("Accounts Matched", f"{n_match} / {n_total}")
        col2.metric("Expected Non-Issues", n_expected)
        col3.metric("Real Mismatches", n_real_mismatch)

        if n_real_mismatch == 0:
            st.success("All accounts reconcile (aside from expected, flagged non-issues).")
        else:
            st.error(f"{n_real_mismatch} account(s) have a real mismatch — review below.")

        show_only_issues = st.checkbox("Show only mismatches", value=(n_real_mismatch > 0))
        display_df = recon_df[recon_df["Checker"] != "Match"] if show_only_issues else recon_df
        st.dataframe(display_df, use_container_width=True)

        csv_data = recon_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download Reconciliation CSV",
            data=csv_data,
            file_name="pnl_reconciliation.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
        )

        st.divider()

        with st.expander(f"Balance Sheet accounts excluded from reconciliation ({len(extra_gl_df)})"):
            st.caption(
                "These GL accounts (Class = SICB Management) have activity but don't appear on the Portfolio "
                "P&L. A Profit & Loss report can only ever contain Income Statement accounts, so these are "
                "Balance Sheet accounts (bank, credit card, inventory, loans, etc.) and are excluded from the "
                "reconciliation by design."
            )
            st.dataframe(extra_gl_df, use_container_width=True)

        st.divider()

        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("← Upload Different Files", use_container_width=True, key="tool5_back_upload"):
                st.session_state.tool5_step = "upload"
                st.rerun()
        with col2:
            if st.button("Restart", use_container_width=True, key="tool5_restart"):
                st.session_state.tool5_step = "upload"
                st.session_state.tool5_recon_df = pd.DataFrame()
                st.session_state.tool5_extra_gl_df = pd.DataFrame()
                st.session_state.tool5_portfolio_raw_list = []
                st.session_state.tool5_property_raw_list = []
                st.session_state.tool5_gl_raw_list = []
                st.session_state.tool5_unit_econ_raw_df = pd.DataFrame()
                st.session_state.tool5_dept_remap = None
                st.session_state.tool5_unit_econ_df = pd.DataFrame()
                st.session_state.tool5_property_merges = []
                st.session_state.tool5_typo_overrides = {}
                st.session_state.tool5_net_income_check_df = pd.DataFrame()
                st.rerun()
        with col3:
            if st.button("Continue to Data Preparation →", type="primary", use_container_width=True, key="tool5_to_prep"):
                with st.spinner("Extracting Owner/Property/Department from the GL..."):
                    portfolio_dfs, property_dfs = [], []
                    for portfolio_raw, property_raw in zip(
                        st.session_state.tool5_portfolio_raw_list, st.session_state.tool5_property_raw_list
                    ):
                        portfolio_df, property_df = tool5_parse_period_pnls(portfolio_raw, property_raw)
                        portfolio_dfs.append(portfolio_df)
                        property_dfs.append(property_df)

                    expanded_accounts, account_section, dup_account_sections = tool5_build_account_universe(
                        portfolio_dfs, property_dfs
                    )

                    unit_econ_pieces = [
                        tool5_extract_unit_economics(gl_raw, expanded_accounts, account_section, dup_account_sections)
                        for gl_raw in st.session_state.tool5_gl_raw_list
                    ]
                    unit_econ_raw = pd.concat(unit_econ_pieces, ignore_index=True)
                st.session_state.tool5_unit_econ_raw_df = unit_econ_raw
                st.session_state.tool5_dept_remap = None
                st.session_state.tool5_step = "choose_prep_path"
                st.rerun()

    # ── STEP: CHOOSE DATA PREP PATH ───────────────────────────────────────────

    elif st.session_state.tool5_step == "choose_prep_path":

        st.subheader("What do you want to prepare?")
        st.caption("The reconciled GL data is ready. Choose which downstream dataset to prepare next.")

        card_style = """
            <style>
            div[data-testid="stVerticalBlock"] div[data-testid="stVerticalBlockBorderWrapper"] {
                height: 100%;
            }
            </style>
        """
        st.markdown(card_style, unsafe_allow_html=True)

        PREP_PATHS = [
            {
                "key": "tool5_path_unit_econ",
                "title": "Property Level P&L",
                "subtitle": "Unit Economics",
                "description": "Extract Owner/Property/Department detail from the GL and prepare a clean, per-property P&L dataset.",
                "enabled": True,
                "next_step": "prep_department",
            },
            {
                "key": "tool5_path_fin_stmt",
                "title": "Financial Statement",
                "subtitle": "Consolidated Portfolio P&L",
                "description": "Consolidate the uploaded periods' Portfolio P&L into one flat, aligned export, ready for your mapping workbook.",
                "enabled": True,
                "next_step": "fin_stmt_export",
            },
            {
                "key": "tool5_path_corp_dash",
                "title": "Corporate Expenses Dashboard",
                "subtitle": "Coming soon",
                "description": "Prepare data for the corporate expenses dashboard.",
                "enabled": False,
                "next_step": None,
            },
        ]

        cols = st.columns(3, gap="large")
        for col, path in zip(cols, PREP_PATHS):
            with col:
                with st.container(border=True):
                    st.markdown(f"#### {path['title']}")
                    st.caption(path["subtitle"])
                    st.write(path["description"])
                    if st.button(
                        "Select" if path["enabled"] else "Coming Soon",
                        key=path["key"],
                        use_container_width=True,
                        type="primary" if path["enabled"] else "secondary",
                        disabled=not path["enabled"],
                    ):
                        st.session_state.tool5_step = path["next_step"]
                        st.rerun()

        st.divider()
        if st.button("← Back to Reconciliation", key="tool5_choose_path_back"):
            st.session_state.tool5_step = "reconcile"
            st.rerun()

    # ── STEP: FINANCIAL STATEMENT — CONSOLIDATED PORTFOLIO P&L ───────────────

    elif st.session_state.tool5_step == "fin_stmt_export":

        st.subheader("Consolidated Portfolio P&L")
        st.caption(
            "Every row from each uploaded period's Portfolio P&L — leaf accounts and subtotal/rollup rows alike "
            "(Total ..., Gross Profit, Net Ordinary Income, Net Other Income, Net Income) — aligned into one flat "
            "table. An account that only appears in some periods still gets its own row, blank for the periods "
            "where it's absent. This is the raw account-level data, not yet mapped to your Financial Statement "
            "line categories — bring it into your mapping workbook as the source for VLOOKUP/SUMIFS."
        )

        portfolio_raw_list = st.session_state.tool5_portfolio_raw_list
        n_periods = len(portfolio_raw_list)

        st.write("**Label each period** (used as the column header below):")
        period_cols = st.columns(n_periods)
        period_labels = []
        for i, col in enumerate(period_cols, start=1):
            with col:
                default_label = f"Period {i}"
                label = st.text_input(f"Period {i}", value=default_label, key=f"tool5_fin_stmt_label_{i}")
                period_labels.append(label.strip() if label.strip() else default_label)

        if len(set(period_labels)) != len(period_labels):
            st.error("Period labels must be unique.")
        else:
            consolidated_df = tool5_build_consolidated_portfolio_pnl(portfolio_raw_list, period_labels)

            st.divider()
            st.success(f"{len(consolidated_df):,} rows across {n_periods} period(s).")
            st.dataframe(
                consolidated_df[["Section", "Account"] + period_labels],
                use_container_width=True,
            )

            excel_bytes = tool5_export_consolidated_pnl_excel(consolidated_df, period_labels)
            st.download_button(
                label="Download Consolidated P&L (Excel)",
                data=excel_bytes,
                file_name="hps_consolidated_portfolio_pnl.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        st.divider()
        if st.button("← Back", key="tool5_fin_stmt_back"):
            st.session_state.tool5_step = "choose_prep_path"
            st.rerun()

    # ── STEP: UNIT ECONOMICS — DEPARTMENT CLEANUP ─────────────────────────────

    elif st.session_state.tool5_step == "prep_department":

        st.subheader("Step 2 — Prepare Data for P&L by Property (Unit Economics)")
        st.caption(
            "Owner and Property have been extracted from the GL's Name field (\"Owner -C:Property\"). "
            "Transactions with no colon in Name (e.g. \"SICB - Rent\", \"SICB - Customer\") are unattributed "
            "and bucketed as Corporate for both Owner and Property."
        )

        raw_df = st.session_state.tool5_unit_econ_raw_df
        st.info(f"{len(raw_df):,} relevant transactions extracted (Class = SICB Management, P&L accounts only).")

        st.divider()
        st.write("**Review Department cleanup**")
        st.caption(
            "Each raw Department value found in the data is shown below, pre-filled with what the default "
            "logic produces (text after the first \":\", or unchanged if there's no colon). Edit any value "
            "you want to override."
        )

        raw_departments = sorted(raw_df["Department (Raw)"].dropna().unique().tolist())

        if st.session_state.tool5_dept_remap is None:
            working_remap = {d: tool5_default_department(d) for d in raw_departments}
        else:
            working_remap = dict(st.session_state.tool5_dept_remap)
            for d in raw_departments:
                if d not in working_remap:
                    working_remap[d] = tool5_default_department(d)

        if not raw_departments:
            st.info("No Department values found in the relevant transactions.")
        else:
            for i, dept in enumerate(raw_departments):
                key = f"tool5_dept_input_{i}"
                if key not in st.session_state:
                    st.session_state[key] = working_remap[dept]

            cols = st.columns(2)
            for i, dept in enumerate(raw_departments):
                with cols[i % 2]:
                    st.text_input(f"`{dept}`", key=f"tool5_dept_input_{i}")

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Reconciliation", use_container_width=True, key="tool5_dept_back"):
                st.session_state.tool5_step = "reconcile"
                st.rerun()
        with col2:
            if st.button("Apply & Continue →", type="primary", use_container_width=True, key="tool5_dept_apply"):
                new_remap = {}
                for i, dept in enumerate(raw_departments):
                    new_value = st.session_state.get(f"tool5_dept_input_{i}", dept)
                    new_remap[dept] = new_value.strip() if new_value.strip() else dept
                st.session_state.tool5_dept_remap = new_remap

                final_df = raw_df.copy()
                final_df["Department"] = final_df["Department (Raw)"].map(
                    lambda d: new_remap.get(d, tool5_default_department(d)) if pd.notna(d) else d
                )
                final_df = final_df.drop(columns=["Department (Raw)"])
                st.session_state.tool5_unit_econ_df = final_df
                for k in [k for k in st.session_state if k.startswith("tool5_dept_input_")]:
                    del st.session_state[k]
                st.session_state.tool5_step = "merge_properties"
                st.rerun()

    # ── STEP: MERGE PROPERTIES ────────────────────────────────────────────────

    elif st.session_state.tool5_step == "merge_properties":

        st.subheader("Step 3 — Merge Properties")
        st.caption("Group property names that refer to the same property. Select all variants, then pick which name to keep.")

        unit_econ_df = st.session_state.tool5_unit_econ_df
        all_properties = sorted(unit_econ_df["Property"].unique().tolist())

        if st.session_state.tool5_property_merges:
            st.write("**Current merge groups:**")
            for i, group in enumerate(st.session_state.tool5_property_merges):
                col1, col2 = st.columns([6, 1])
                with col1:
                    variants_str = ", ".join(f"`{v}`" for v in group["variants"] if v != group["canonical"])
                    st.write(f"{variants_str} → **{group['canonical']}**")
                with col2:
                    if st.button("Remove", key=f"tool5_remove_merge_{i}"):
                        st.session_state.tool5_property_merges.pop(i)
                        st.rerun()
            st.divider()

        st.write("**Add a new merge group:**")
        n = len(st.session_state.tool5_property_merges)
        selected_variants = st.multiselect(
            "Select property names to merge",
            options=all_properties,
            key=f"tool5_merge_variants_{n}",
        )

        if len(selected_variants) >= 2:
            canonical = st.radio(
                "Which name to keep?",
                options=selected_variants,
                key=f"tool5_merge_canonical_{n}",
            )
            if st.button("Add Group", type="primary", key="tool5_merge_add_group"):
                st.session_state.tool5_property_merges.append({
                    "variants": selected_variants,
                    "canonical": canonical,
                })
                st.rerun()

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Department Cleanup", use_container_width=True, key="tool5_merge_back"):
                st.session_state.tool5_step = "prep_department"
                st.rerun()
        with col2:
            if st.button("Apply & Continue →", type="primary", use_container_width=True, key="tool5_merge_apply"):
                st.session_state.tool5_unit_econ_df = tool5_apply_property_merges(
                    unit_econ_df, st.session_state.tool5_property_merges
                )
                st.session_state.tool5_step = "typo_fix"
                st.rerun()

    # ── STEP: FIX PROPERTY/OWNER TYPOS ────────────────────────────────────────

    elif st.session_state.tool5_step == "typo_fix":

        st.subheader("Step 4 — Fix Property/Owner Typos")
        st.caption(
            "Search for a property and correct spelling in its Property name and/or Owner. Each correction "
            "applies to every transaction for that property, across all uploaded periods."
        )

        unit_econ_df = st.session_state.tool5_unit_econ_df
        all_properties = sorted(unit_econ_df["Property"].unique().tolist())

        if st.session_state.tool5_typo_overrides:
            st.write("**Current corrections:**")
            for prop, vals in list(st.session_state.tool5_typo_overrides.items()):
                col1, col2 = st.columns([6, 1])
                with col1:
                    parts = []
                    if "Property" in vals:
                        parts.append(f"Property → **{vals['Property']}**")
                    if "Owner" in vals:
                        parts.append(f"Owner → **{vals['Owner']}**")
                    st.write(f"`{prop}`: " + ", ".join(parts))
                with col2:
                    if st.button("Remove", key=f"tool5_typo_remove_{prop}"):
                        del st.session_state.tool5_typo_overrides[prop]
                        st.rerun()
            st.divider()

        st.write("**Search for a property to correct:**")
        selected_property = st.selectbox("Property", options=all_properties, key="tool5_typo_search")

        if selected_property:
            current_owner_series = unit_econ_df.loc[unit_econ_df["Property"] == selected_property, "Owner"]
            current_owner = current_owner_series.iloc[0] if len(current_owner_series) else ""
            pending = st.session_state.tool5_typo_overrides.get(selected_property, {})

            col1, col2 = st.columns(2)
            with col1:
                new_property = st.text_input(
                    "Corrected Property name",
                    value=pending.get("Property", selected_property),
                    key=f"tool5_typo_property_input_{selected_property}",
                )
            with col2:
                new_owner = st.text_input(
                    "Corrected Owner name",
                    value=pending.get("Owner", current_owner),
                    key=f"tool5_typo_owner_input_{selected_property}",
                )

            if st.button("Save Correction", type="primary", key="tool5_typo_save"):
                entry = {}
                if new_property.strip() and new_property.strip() != selected_property:
                    entry["Property"] = new_property.strip()
                if new_owner.strip() and new_owner.strip() != current_owner:
                    entry["Owner"] = new_owner.strip()
                if entry:
                    st.session_state.tool5_typo_overrides[selected_property] = entry
                else:
                    st.session_state.tool5_typo_overrides.pop(selected_property, None)
                st.rerun()

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Merge Properties", use_container_width=True, key="tool5_typo_back"):
                st.session_state.tool5_step = "merge_properties"
                st.rerun()
        with col2:
            if st.button("Apply & Continue →", type="primary", use_container_width=True, key="tool5_typo_apply"):
                df = st.session_state.tool5_unit_econ_df.copy()
                for prop, vals in st.session_state.tool5_typo_overrides.items():
                    mask = df["Property"] == prop
                    if "Property" in vals:
                        df.loc[mask, "Property"] = vals["Property"]
                    if "Owner" in vals:
                        df.loc[mask, "Owner"] = vals["Owner"]
                st.session_state.tool5_unit_econ_df = df
                st.session_state.tool5_step = "net_income_check"
                st.rerun()

    # ── STEP: PER-PROPERTY NET INCOME CHECK (FINAL RECONCILIATION) ────────────

    elif st.session_state.tool5_step == "net_income_check":

        st.subheader("Step 5 — Final Reconciliation: Per-Property Net Income")
        st.caption(
            "For each property, Net Income derived from the extracted (merged, Corporate-filtered) GL data "
            "is compared against the Net Income shown for that property on the Property-Level P&L."
        )

        check_df = tool5_build_net_income_check(
            st.session_state.tool5_unit_econ_df,
            st.session_state.tool5_property_raw_list,
            st.session_state.tool5_property_merges,
        )
        st.session_state.tool5_net_income_check_df = check_df

        n_total = len(check_df)
        n_match = int(check_df["Match"].sum())
        col1, col2 = st.columns(2)
        col1.metric("Properties Matched", f"{n_match} / {n_total}")
        col2.metric("Mismatches", n_total - n_match)

        if n_match == n_total:
            st.success("All properties reconcile — Net Income matches the Property-Level P&L for every property.")
        else:
            st.error(f"{n_total - n_match} propert(y/ies) do not reconcile — review below.")

        show_only_issues = st.checkbox("Show only mismatches", value=(n_match < n_total))
        display_df = check_df[~check_df["Match"]] if show_only_issues else check_df

        def _highlight_match(row):
            color = "background-color: #d4f7d4" if row["Match"] else ""
            return [color] * len(row)

        st.dataframe(
            display_df.style.apply(_highlight_match, axis=1).format(
                {"Property P&L Net Income": "{:,.2f}", "GL Net Income": "{:,.2f}", "Diff": "{:,.2f}"}
            ),
            use_container_width=True,
        )

        st.divider()
        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Fix Typos", use_container_width=True, key="tool5_netcheck_back"):
                st.session_state.tool5_step = "typo_fix"
                st.rerun()
        with col2:
            if st.button("Continue to Export →", type="primary", use_container_width=True, key="tool5_netcheck_continue"):
                st.session_state.tool5_step = "prep_export"
                st.rerun()

    # ── STEP: UNIT ECONOMICS — EXPORT ─────────────────────────────────────────

    elif st.session_state.tool5_step == "prep_export":

        final_df = st.session_state.tool5_unit_econ_df
        raw_df = st.session_state.tool5_unit_econ_raw_df

        raw_sum = raw_df["Amount"].sum(min_count=1)
        final_sum = final_df["Amount"].sum(min_count=1)
        if abs(raw_sum - final_sum) < 0.01:
            st.success(f"Reconciliation passed — Amount totals match: {raw_sum:,.2f}")
        else:
            st.error(
                f"Reconciliation failed — Extracted total: {raw_sum:,.2f} | "
                f"Final total: {final_sum:,.2f} | Difference: {raw_sum - final_sum:,.2f}"
            )

        n_properties = final_df.loc[final_df["Property"] != "Corporate", "Property"].nunique()
        n_corporate_rows = int((final_df["Property"] == "Corporate").sum())
        col1, col2, col3 = st.columns(3)
        col1.metric("Rows", f"{len(final_df):,}")
        col2.metric("Distinct Properties", n_properties)
        col3.metric("Corporate (Unattributed) Rows", f"{n_corporate_rows:,}")

        TOOL5_PREVIEW_ROWS = 200
        st.caption(f"Showing the first {TOOL5_PREVIEW_ROWS} of {len(final_df):,} rows — download the CSVs below for the full data.")
        st.dataframe(final_df.head(TOOL5_PREVIEW_ROWS), use_container_width=True)

        st.divider()

        portfolio_net_income = tool5_portfolio_net_income(st.session_state.tool5_portfolio_raw_list)
        # Net Income is computed once from the pre-sign-flip filtered data (same
        # -sum(raw Amount) identity used everywhere else) — both exports share
        # this same figure regardless of grouping, since it's the same filtered rows.
        combined_net_income = -tool5_apply_corporate_filter(final_df)["Amount"].sum()

        def _show_export_net_income():
            diff = portfolio_net_income - combined_net_income
            if abs(diff) < TOOL5_TOLERANCE:
                st.success(f"Net Income from this export: {combined_net_income:,.2f} — matches the Portfolio P&L.")
            else:
                st.error(
                    f"Net Income from this export: {combined_net_income:,.2f} — does NOT match the Portfolio "
                    f"P&L ({portfolio_net_income:,.2f}), difference: {diff:,.2f}."
                )

        st.subheader("Export 1 — Property-Level P&L (Grouped)")
        st.caption(
            "Grouped by Accounting Period + Account + Property, summed Amount. Corporate transactions are "
            "excluded, except for accounts below Net Ordinary Income (Other Income / Other Expense) — and "
            "even there, the two OH-split accounts are still excluded since their Corporate rows are just "
            "the offsetting reclass entry, not real unattributed money. Amount's sign is flipped for "
            "Income/Other Income accounts so revenue displays as positive, matching the P&L reports."
        )
        export1_df = tool5_build_export_grouped(final_df)
        _show_export_net_income()
        st.caption(f"Showing the first {TOOL5_PREVIEW_ROWS} of {len(export1_df):,} rows — download the CSV below for the full data.")
        st.dataframe(export1_df.head(TOOL5_PREVIEW_ROWS), use_container_width=True)
        export1_csv = tool5_to_export_csv(
            export1_df, text_columns=["Property", "Owner", "Property Owner Type"]
        )
        st.download_button(
            label="Download Export 1 — Property-Level P&L CSV",
            data=export1_csv,
            file_name="property_level_pnl.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
            key="tool5_download_export1",
        )

        st.divider()

        st.subheader("Export 2 — Transaction Detail (Ungrouped)")
        st.caption(
            "Same Corporate-exclusion filter and sign convention as Export 1, but one row per raw GL "
            "transaction, with full transaction detail retained (Date, Type, Memo, Name, etc.)."
        )
        export2_df = tool5_build_export_detail(final_df)
        _show_export_net_income()
        st.caption(f"Showing the first {TOOL5_PREVIEW_ROWS} of {len(export2_df):,} rows — download the CSV below for the full data.")
        st.dataframe(export2_df.head(TOOL5_PREVIEW_ROWS), use_container_width=True)
        export2_csv = tool5_to_export_csv(
            export2_df, text_columns=["Property", "Owner", "Property Owner Type"]
        )
        st.download_button(
            label="Download Export 2 — Transaction Detail CSV",
            data=export2_csv,
            file_name="unit_economics_detail.csv",
            mime="text/csv",
            type="primary",
            use_container_width=True,
            key="tool5_download_export2",
        )

        st.divider()

        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Back to Final Reconciliation", use_container_width=True, key="tool5_export_back"):
                st.session_state.tool5_step = "net_income_check"
                st.rerun()
        with col2:
            if st.button("Restart", use_container_width=True, key="tool5_export_restart"):
                st.session_state.tool5_step = "upload"
                st.session_state.tool5_recon_df = pd.DataFrame()
                st.session_state.tool5_extra_gl_df = pd.DataFrame()
                st.session_state.tool5_portfolio_raw_list = []
                st.session_state.tool5_property_raw_list = []
                st.session_state.tool5_gl_raw_list = []
                st.session_state.tool5_unit_econ_raw_df = pd.DataFrame()
                st.session_state.tool5_dept_remap = None
                st.session_state.tool5_unit_econ_df = pd.DataFrame()
                st.session_state.tool5_property_merges = []
                st.session_state.tool5_net_income_check_df = pd.DataFrame()
                for k in [k for k in st.session_state if k.startswith("tool5_dept_input_")]:
                    del st.session_state[k]
                st.rerun()
